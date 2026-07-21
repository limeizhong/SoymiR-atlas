#!/usr/bin/env python3
"""Build Supplementary Table S3 tissue expression profiles.

Input expression rows are expected to contain:
sample, mature_seq, Tissue, reads, Seq-ID, Reported_Status, Total_Reads.

For each library, read abundance is normalized as RPM. Tissue-level
expression is the average RPM across libraries in which the miRNA is detected
within that tissue. S3 values are log2(mean RPM + 1), rounded to two decimals.
"""

from __future__ import annotations

import argparse
import copy
import math
import re
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl


SCRIPT_DIR = Path(__file__).resolve().parent
MODULE_DIR = SCRIPT_DIR.parent
INPUT_DIR = MODULE_DIR / "input"
RESULTS_DIR = MODULE_DIR / "results"

DEFAULT_EXPRESSION = INPUT_DIR / "68282.1588.expression.rawdata.txt"
DEFAULT_TEMPLATE = INPUT_DIR / "2026_merged_supplementary_tables_template.xlsx"
DEFAULT_S3_TSV = RESULTS_DIR / "supplementary_table_S3_tissue_expression_profiles.tsv"
DEFAULT_WORKBOOK = RESULTS_DIR / "2026_merged_supplementary_tables_S3_updated.xlsx"

TISSUES = [
    "cotyledon",
    "flower",
    "leaves",
    "mixed",
    "nodules",
    "pod",
    "root",
    "seed",
    "shoot",
    "stem",
]

READ_COUNT_RE = re.compile(r"_x(\d+)$")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--expression", type=Path, default=DEFAULT_EXPRESSION)
    parser.add_argument("--template-workbook", type=Path, default=DEFAULT_TEMPLATE)
    parser.add_argument("--s3-tsv", type=Path, default=DEFAULT_S3_TSV)
    parser.add_argument("--output-workbook", type=Path, default=DEFAULT_WORKBOOK)
    return parser.parse_args()


def parse_read_count(read_id: str) -> int:
    match = READ_COUNT_RE.search(read_id)
    if not match:
        raise ValueError(f"Cannot parse read count from reads field: {read_id}")
    return int(match.group(1))


def load_expression(expression_file: Path):
    sample_tissue: dict[str, str] = {}
    sample_total_reads: dict[str, int] = {}
    sample_mirna_counts: defaultdict[tuple[str, str], int] = defaultdict(int)
    mirna_ids: set[str] = set()
    seen_read_rows: set[tuple[str, str, str]] = set()
    duplicate_rows = 0
    input_rows = 0

    with expression_file.open() as handle:
        header = handle.readline().rstrip("\n").split("\t")
        col = {name: idx for idx, name in enumerate(header)}
        required = ["sample", "mature_seq", "Tissue", "reads", "Seq-ID", "Total_Reads"]
        missing = [name for name in required if name not in col]
        if missing:
            raise ValueError(f"Missing required columns in {expression_file}: {missing}")

        for line_no, line in enumerate(handle, start=2):
            fields = line.rstrip("\n").split("\t")
            if len(fields) != len(header):
                raise ValueError(f"Unexpected column count at line {line_no}: {len(fields)}")

            input_rows += 1
            sample = fields[col["sample"]]
            mature_seq = fields[col["mature_seq"]]
            tissue = fields[col["Tissue"]].strip().lower()
            read_id = fields[col["reads"]]
            mirna_id = fields[col["Seq-ID"]]
            total_reads = int(fields[col["Total_Reads"]])

            if tissue not in TISSUES:
                raise ValueError(f"Unexpected tissue {tissue!r} at line {line_no}")
            if sample in sample_tissue and sample_tissue[sample] != tissue:
                raise ValueError(f"Sample {sample} has conflicting tissue labels")
            if sample in sample_total_reads and sample_total_reads[sample] != total_reads:
                raise ValueError(f"Sample {sample} has conflicting Total_Reads")

            sample_tissue[sample] = tissue
            sample_total_reads[sample] = total_reads
            mirna_ids.add(mirna_id)

            dedup_key = (sample, mature_seq, read_id)
            if dedup_key in seen_read_rows:
                duplicate_rows += 1
                continue
            seen_read_rows.add(dedup_key)
            sample_mirna_counts[(sample, mirna_id)] += parse_read_count(read_id)

    samples_by_tissue = {tissue: [] for tissue in TISSUES}
    for sample, tissue in sample_tissue.items():
        samples_by_tissue[tissue].append(sample)
    for tissue in TISSUES:
        samples_by_tissue[tissue].sort()
        if not samples_by_tissue[tissue]:
            raise ValueError(f"No libraries assigned to tissue {tissue}")

    return {
        "input_rows": input_rows,
        "duplicate_rows": duplicate_rows,
        "sample_tissue": sample_tissue,
        "sample_total_reads": sample_total_reads,
        "sample_mirna_counts": sample_mirna_counts,
        "mirna_ids": sorted(mirna_ids),
        "samples_by_tissue": samples_by_tissue,
    }


def build_s3_rows(expression_data) -> list[list[object]]:
    sample_total_reads = expression_data["sample_total_reads"]
    sample_mirna_counts = expression_data["sample_mirna_counts"]
    samples_by_tissue = expression_data["samples_by_tissue"]

    rows: list[list[object]] = []
    for mirna_id in expression_data["mirna_ids"]:
        values = []
        for tissue in TISSUES:
            rpm_values = []
            for sample in samples_by_tissue[tissue]:
                count = sample_mirna_counts.get((sample, mirna_id))
                if count:
                    rpm_values.append(count / sample_total_reads[sample] * 1_000_000)
            mean_rpm = sum(rpm_values) / len(rpm_values) if rpm_values else 0
            xi = math.log2(mean_rpm + 1) if mean_rpm > 0 else 0
            values.append(round(xi, 2))
        rows.append([mirna_id, *values])
    return rows


def write_s3_tsv(rows: list[list[object]], output_file: Path) -> None:
    output_file.parent.mkdir(parents=True, exist_ok=True)
    with output_file.open("w") as handle:
        handle.write("\t".join(["miRNA_ID", *TISSUES]) + "\n")
        for row in rows:
            handle.write("\t".join(str(value) for value in row) + "\n")


def copy_cell_style(source, target) -> None:
    target.font = copy.copy(source.font)
    target.fill = copy.copy(source.fill)
    target.border = copy.copy(source.border)
    target.alignment = copy.copy(source.alignment)
    target.number_format = source.number_format
    target.protection = copy.copy(source.protection)


def update_workbook_s3(template_file: Path, rows: list[list[object]], output_file: Path) -> None:
    output_file.parent.mkdir(parents=True, exist_ok=True)
    workbook = openpyxl.load_workbook(template_file)
    worksheet = workbook["S3"]

    header_styles = [copy.copy(worksheet.cell(2, col)) for col in range(1, 12)]
    data_styles = [copy.copy(worksheet.cell(3, col)) for col in range(1, 12)]

    if worksheet.max_row > 2:
        worksheet.delete_rows(3, worksheet.max_row - 2)

    worksheet.cell(1, 1).value = (
        "Supplementary Table 3. Tissue expression profiles of 1,588 "
        "non-redundant mature soybean miRNAs across 10 major tissue categories."
    )

    for col, value in enumerate(["miRNA_ID", *TISSUES], start=1):
        cell = worksheet.cell(2, col)
        cell.value = value
        copy_cell_style(header_styles[col - 1], cell)

    for row_index, row in enumerate(rows, start=3):
        for col_index, value in enumerate(row, start=1):
            cell = worksheet.cell(row_index, col_index)
            cell.value = value
            copy_cell_style(data_styles[col_index - 1], cell)

    worksheet.auto_filter.ref = f"A2:K{len(rows) + 2}"
    workbook.save(output_file)


def main() -> None:
    args = parse_args()
    expression_data = load_expression(args.expression)
    rows = build_s3_rows(expression_data)
    write_s3_tsv(rows, args.s3_tsv)
    update_workbook_s3(args.template_workbook, rows, args.output_workbook)

    print(args.s3_tsv)
    print(args.output_workbook)
    print(f"Input_rows: {expression_data['input_rows']}")
    print(f"Duplicate_rows_skipped: {expression_data['duplicate_rows']}")
    print(f"Libraries: {len(expression_data['sample_tissue'])}")
    print(f"miRNA_IDs: {len(expression_data['mirna_ids'])}")
    print("Libraries_by_tissue:")
    for tissue in TISSUES:
        print(f"{tissue}: {len(expression_data['samples_by_tissue'][tissue])}")


if __name__ == "__main__":
    main()
