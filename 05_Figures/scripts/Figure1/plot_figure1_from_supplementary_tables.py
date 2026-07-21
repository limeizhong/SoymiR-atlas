#!/usr/bin/env python3
from collections import Counter, defaultdict
import os
import re
from pathlib import Path

SCRIPT_DIR = Path(__file__).resolve().parent
MODULE_DIR = SCRIPT_DIR.parents[1]
os.environ.setdefault("MPLCONFIGDIR", str(MODULE_DIR / "results/.matplotlib_cache"))
os.environ.setdefault("XDG_CACHE_HOME", str(MODULE_DIR / "results/.cache"))
(MODULE_DIR / "results/.matplotlib_cache").mkdir(parents=True, exist_ok=True)
(MODULE_DIR / "results/.cache").mkdir(parents=True, exist_ok=True)

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.ticker import MaxNLocator
from openpyxl import load_workbook


WORKBOOK = MODULE_DIR / "input/source_tables/2026_merged_supplementary_tables.xlsx"
OUT = MODULE_DIR / "results/final_figures/Figure_1.png"

BLUE = "#0879B5"
LIGHT_BLUE = "#5AB4E6"


def add_labels(ax, bars, offset=0.02, fontsize=15):
    ymin, ymax = ax.get_ylim()
    span = ymax - ymin
    for bar in bars:
        h = bar.get_height()
        ax.text(
            bar.get_x() + bar.get_width() / 2,
            h + span * offset,
            f"{int(h):,}",
            ha="center",
            va="bottom",
            fontsize=fontsize,
        )


def clean_axes(ax):
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["left"].set_linewidth(1.6)
    ax.spines["bottom"].set_linewidth(1.6)
    ax.tick_params(axis="both", width=1.6, length=5, labelsize=15)
    ax.yaxis.set_major_locator(MaxNLocator(integer=True))


def panel_label(ax, label):
    ax.text(
        -0.15,
        1.08,
        label,
        transform=ax.transAxes,
        fontsize=24,
        fontweight="bold",
        ha="left",
        va="center",
    )


def get_r(seq_id):
    m = re.search(r"_R(\d+)_", str(seq_id))
    return int(m.group(1)) if m else None


def main():
    wb = load_workbook(WORKBOOK, read_only=True, data_only=True)

    # A. Tissue distribution of sRNA-seq libraries.
    ws = wb["S1"]
    h = [c.value for c in ws[2]]
    tissue_idx = h.index("Tissue Catagory")
    tissue_counts = Counter()
    for row in ws.iter_rows(min_row=3, values_only=True):
        tissue = row[tissue_idx]
        if not tissue:
            continue
        tissue = str(tissue).strip().lower()
        tissue_counts[tissue] += 1
    tissue_order = sorted(tissue_counts, key=lambda x: (-tissue_counts[x], x))
    tissue_labels = [t.capitalize() for t in tissue_order]
    tissue_values = [tissue_counts[t] for t in tissue_order]

    # S2 data.
    ws = wb["S2"]
    h = [c.value for c in ws[2]]
    s2 = [dict(zip(h, row)) for row in ws.iter_rows(min_row=3, values_only=True) if row[0] is not None]
    seq_by_id = {}
    for row in s2:
        sid = row["Seq-ID"]
        seq = row["Sequences_Mature"]
        if sid and seq:
            seq_by_id[sid] = str(seq)

    # B. Mature miRNA length distribution.
    length_counts = Counter(len(seq) for seq in seq_by_id.values())
    length_order = [20, 21, 22, 23, 24]
    length_values = [length_counts.get(x, 0) for x in length_order]

    # C. Library detection coverage based on R in Seq-ID.
    coverage_bins = ["2-5", "6-10", "11-50", "51-100", ">100"]
    coverage_counts = dict.fromkeys(coverage_bins, 0)
    for sid in seq_by_id:
        r = get_r(sid)
        if r is None:
            continue
        if 2 <= r <= 5:
            coverage_counts["2-5"] += 1
        elif 6 <= r <= 10:
            coverage_counts["6-10"] += 1
        elif 11 <= r <= 50:
            coverage_counts["11-50"] += 1
        elif 51 <= r <= 100:
            coverage_counts["51-100"] += 1
        elif r > 100:
            coverage_counts[">100"] += 1
    coverage_values = [coverage_counts[b] for b in coverage_bins]

    # D. Tissue expression breadth.
    ws = wb["S3"]
    h = [c.value for c in ws[2]]
    breadth_counts = Counter()
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row[0]:
            continue
        n = sum(1 for v in row[1:] if v not in (None, "") and float(v) > 0)
        breadth_counts[n] += 1
    breadth_order = list(range(1, 11))
    breadth_values = [breadth_counts.get(x, 0) for x in breadth_order]

    # E. Annotation resources from S2.
    precursor_total = len(s2)
    mature_total = len(seq_by_id)
    resource_labels = ["SoymiR", "miRBase", "PmiREN"]
    precursor_values = [
        precursor_total,
        756,
        1349,
    ]
    mature_values = [
        mature_total,
        569,
        840,
    ]

    # F. Family size distribution from S2 (count rows per Family).
    ws = wb["S2"]
    h = [c.value for c in ws[2]]
    family_idx = h.index("Family")
    family_row_counts = Counter()
    for row in ws.iter_rows(min_row=3, values_only=True):
        family = row[family_idx]
        if not family:
            continue
        family_row_counts[str(family).strip()] += 1
    family_bins = ["1", "2", "3-5", "6-10", ">10"]
    family_counts = dict.fromkeys(family_bins, 0)
    for cnt in family_row_counts.values():
        if cnt == 1:
            family_counts["1"] += 1
        elif cnt == 2:
            family_counts["2"] += 1
        elif 3 <= cnt <= 5:
            family_counts["3-5"] += 1
        elif 6 <= cnt <= 10:
            family_counts["6-10"] += 1
        elif cnt > 10:
            family_counts[">10"] += 1
    family_values = [family_counts[b] for b in family_bins]

    plt.rcParams.update(
        {
            "font.family": "DejaVu Sans",
            "axes.titleweight": "bold",
            "axes.titlesize": 19,
            "axes.labelsize": 16,
            "xtick.labelsize": 15,
            "ytick.labelsize": 15,
        }
    )

    fig, axs = plt.subplots(3, 2, figsize=(16.4, 18.6), dpi=300)
    fig.subplots_adjust(left=0.09, right=0.98, top=0.96, bottom=0.08, hspace=0.62, wspace=0.36)

    # A
    ax = axs[0, 0]
    bars = ax.bar(range(len(tissue_labels)), tissue_values, color=BLUE, width=0.72)
    ax.set_title("Tissue distribution of sRNA-seq libraries", loc="left", pad=12)
    ax.set_ylabel("Number of libraries")
    ax.set_xticks(range(len(tissue_labels)), tissue_labels, rotation=45, ha="right")
    ax.set_ylim(0, max(tissue_values) * 1.18)
    clean_axes(ax)
    add_labels(ax, bars)
    panel_label(ax, "A")

    # B
    ax = axs[0, 1]
    bars = ax.bar(range(len(length_order)), length_values, color=BLUE, width=0.72)
    ax.set_title("Mature miRNA length distribution", loc="left", pad=12)
    ax.set_ylabel("Number of miRNAs")
    ax.set_xlabel("Mature miRNA sequence length (nt)")
    ax.set_xticks(range(len(length_order)), [str(x) for x in length_order])
    ax.set_ylim(0, max(length_values) * 1.18)
    clean_axes(ax)
    add_labels(ax, bars)
    panel_label(ax, "B")

    # C
    ax = axs[1, 0]
    bars = ax.bar(range(len(coverage_bins)), coverage_values, color=BLUE, width=0.72)
    ax.set_title("Library detection coverage", loc="left", pad=12)
    ax.set_ylabel("Number of miRNAs")
    ax.set_xlabel("Detection coverage level (libraries)")
    ax.set_xticks(range(len(coverage_bins)), coverage_bins)
    ax.set_ylim(0, max(coverage_values) * 1.18)
    clean_axes(ax)
    add_labels(ax, bars)
    panel_label(ax, "C")

    # D
    ax = axs[1, 1]
    bars = ax.bar(range(len(breadth_order)), breadth_values, color=BLUE, width=0.72)
    ax.set_title("Tissue expression breadth", loc="left", pad=12)
    ax.set_ylabel("Number of miRNAs")
    ax.set_xlabel("Number of tissues expressing miRNA")
    ax.set_xticks(range(len(breadth_order)), [str(x) for x in breadth_order])
    ax.set_ylim(0, max(breadth_values) * 1.18)
    clean_axes(ax)
    add_labels(ax, bars)
    panel_label(ax, "D")

    # E
    ax = axs[2, 0]
    x = list(range(len(resource_labels)))
    width = 0.33
    bars1 = ax.bar([i - width / 2 for i in x], precursor_values, color=BLUE, width=width, label="Precursor")
    bars2 = ax.bar([i + width / 2 for i in x], mature_values, color=LIGHT_BLUE, width=width, label="Mature")
    ax.set_title("miRNA annotation resources", loc="left", pad=12)
    ax.set_ylabel("Number of miRNAs")
    ax.set_xlabel("miRNA databases")
    ax.set_xticks(x, resource_labels)
    ax.set_ylim(0, max(precursor_values + mature_values) * 1.24)
    clean_axes(ax)
    add_labels(ax, bars1, fontsize=14)
    add_labels(ax, bars2, fontsize=14)
    ax.legend(frameon=False, loc="upper right", fontsize=14, ncol=2)
    panel_label(ax, "E")

    # F
    ax = axs[2, 1]
    bars = ax.bar(range(len(family_bins)), family_values, color=BLUE, width=0.72)
    ax.set_title("Distribution of miRNA family sizes", loc="left", pad=12)
    ax.set_ylabel("Number of families")
    ax.set_xlabel("Number of members per family")
    ax.set_xticks(range(len(family_bins)), family_bins)
    ax.set_ylim(0, max(family_values) * 1.18)
    clean_axes(ax)
    add_labels(ax, bars)
    panel_label(ax, "F")

    fig.savefig(OUT, dpi=300, facecolor="white")
    print(f"Saved {OUT.resolve()}")
    print("A", dict(zip(tissue_labels, tissue_values)))
    print("B", dict(zip(length_order, length_values)))
    print("C", dict(zip(coverage_bins, coverage_values)))
    print("D", dict(zip(breadth_order, breadth_values)))
    print("E precursor", dict(zip(resource_labels, precursor_values)))
    print("E mature", dict(zip(resource_labels, mature_values)))
    print("F", dict(zip(family_bins, family_values)))


if __name__ == "__main__":
    main()
