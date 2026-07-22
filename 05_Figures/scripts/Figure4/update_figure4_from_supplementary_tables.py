#!/usr/bin/env python3
from __future__ import annotations

from collections import Counter, defaultdict
from pathlib import Path
import csv
import math
import os
import re

import numpy as np
from openpyxl import load_workbook


SCRIPT_DIR = Path(__file__).resolve().parent
ATLAS_ROOT = next((p for p in SCRIPT_DIR.parents if p.name == "SoymiR-atlas"), None)

if ATLAS_ROOT is not None:
    ROOT = ATLAS_ROOT.parent
    WORKBOOK = ATLAS_ROOT / "05_Figures/input/source_tables/2026_merged_supplementary_tables.xlsx"
    DATA_DIR = ATLAS_ROOT / "05_Figures/input/plotting_data/Figure4"
    FIGURE_DIR = ATLAS_ROOT / "05_Figures/results/intermediate_figures/Figure4"
    FINAL_FIGURE_DIR = ATLAS_ROOT / "05_Figures/results/final_figures"
else:
    ROOT = SCRIPT_DIR.parents[4]
    WORKBOOK = ROOT / "2026-miRNA/combine_data/current_database/2026_合并_supplementary_tables.xlsx"
    DATA_DIR = ROOT / "2026-miRNA/figures/working_analysis/data/figure4_targets"
    FIGURE_DIR = ROOT / "2026-miRNA/figures/working_analysis/intermediate_figures/figure4"
    FINAL_FIGURE_DIR = ROOT / "2026-miRNA/figures/manuscript_results_package/01_final_figures"

OUTPUT = FIGURE_DIR / "Figure_miRNA_target_support_count_overlap_ABCD_600dpi.png"
FINAL_OUTPUT = FINAL_FIGURE_DIR / "Figure_4.png"

os.environ.setdefault("MPLCONFIGDIR", str(ROOT / ".matplotlib_cache"))
os.environ.setdefault("XDG_CACHE_HOME", str(ROOT / ".cache"))
(ROOT / ".matplotlib_cache").mkdir(parents=True, exist_ok=True)
(ROOT / ".cache").mkdir(parents=True, exist_ok=True)

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
from scipy.stats import mannwhitneyu


BLUE = "#0072B2"
ORANGE = "#D55E00"
GRAY = "#666666"
SUPPORT_LEVELS = ["1", "2-5", "6-10", "10-20", ">20"]
CONSERVATION_LEVELS = ["Conserved", "Specific"]


def clean(value) -> str:
    return str(value or "").strip()


def is_isomir(row: dict) -> bool:
    variant_status = clean(row.get("variant_status"))
    name = clean(row.get("miRNA_ID"))
    return variant_status == "variant" or bool(re.search(r"\.v\d+$", name))


def canonical_name(miRNA_id: str) -> str:
    return re.sub(r"\.v\d+$", "", clean(miRNA_id))


def to_int(value) -> int:
    return int(float(value))


def midpoint(row: dict, start_key: str, end_key: str) -> float:
    return (to_int(row[start_key]) + to_int(row[end_key])) / 2


def conservation_label(value: str) -> str:
    value = clean(value).lower()
    return "Conserved" if value == "conserved" else "Specific"


def support_class(library_count: int) -> str:
    if library_count == 1:
        return "1"
    if library_count <= 5:
        return "2-5"
    if library_count <= 10:
        return "6-10"
    if library_count <= 20:
        return "10-20"
    return ">20"


def read_sheet(workbook, sheet_name: str) -> list[dict]:
    ws = workbook[sheet_name]
    header = [cell.value for cell in ws[2]]
    rows = []
    for values in ws.iter_rows(min_row=3, values_only=True):
        if not values or values[0] is None:
            continue
        rows.append(dict(zip(header, values)))
    return rows


def choose_detected_canonical(isomir: dict, candidates: list[dict]) -> dict | None:
    if not candidates:
        return None

    def rank(candidate: dict):
        same_locus = candidate["miRNA_loci"] == isomir["miRNA_loci"]
        same_chr_strand = candidate["Chr"] == isomir["Chr"] and candidate["Strand"] == isomir["Strand"]
        same_hairpin = (
            same_chr_strand
            and candidate["H_start"] == isomir["H_start"]
            and candidate["H_end"] == isomir["H_end"]
        )
        return (
            0 if same_locus and same_hairpin else 1 if same_locus and same_chr_strand else 2,
            abs(midpoint(candidate, "M_start", "M_end") - midpoint(isomir, "M_start", "M_end")),
            abs(midpoint(candidate, "H_start", "H_end") - midpoint(isomir, "H_start", "H_end")),
        )

    return min(candidates, key=rank)


def coordinate_offsets(canonical: dict, isomir: dict) -> tuple[int, int]:
    can_start = to_int(canonical["M_start"])
    can_end = to_int(canonical["M_end"])
    iso_start = to_int(isomir["M_start"])
    iso_end = to_int(isomir["M_end"])
    if canonical["Strand"] == "+":
        return iso_start - can_start, iso_end - can_end
    return can_end - iso_end, can_start - iso_start


def sequence_comparison(canonical: dict, isomir: dict) -> dict:
    can = clean(canonical["Sequences_Mature"]).upper().replace("U", "T")
    iso = clean(isomir["Sequences_Mature"]).upper().replace("U", "T")
    offset_5p, offset_3p = coordinate_offsets(canonical, isomir)
    can_start = max(0, offset_5p)
    iso_start = max(0, -offset_5p)
    overlap = min(len(can) - can_start, len(iso) - iso_start)
    mismatches = 0
    if overlap > 0:
        mismatches = sum(can[can_start + i] != iso[iso_start + i] for i in range(overlap))
    return {
        "seed_changed": "yes" if can[1:8] != iso[1:8] else "no",
        "terminal_change_5p": -offset_5p,
        "terminal_change_3p": offset_3p,
        "overlap_length": overlap,
        "overlap_mismatches": mismatches,
    }


def build_canonical_isomir_pairs(s2_rows: list[dict]) -> list[dict]:
    canonical_by_id = defaultdict(list)
    isomirs = []
    for row in s2_rows:
        row["Group"] = "isomiR" if is_isomir(row) else "canonical"
        if row["Group"] == "canonical":
            canonical_by_id[row["miRNA_ID"]].append(row)
        else:
            isomirs.append(row)

    pairs = []
    for row in sorted(isomirs, key=lambda r: to_int(r["No."])):
        expected_id = canonical_name(row["miRNA_ID"])
        canonical = choose_detected_canonical(row, canonical_by_id.get(expected_id, []))
        if canonical is None:
            continue
        comparison = sequence_comparison(canonical, row)
        overlap = comparison["overlap_length"]
        identity = (overlap - comparison["overlap_mismatches"]) / overlap if overlap > 0 else 0
        if overlap < 10 or identity < 0.70:
            continue
        pairs.append({
            "miRNA_loci": row["miRNA_loci"],
            "Family": row["Family"],
            "canonical_miRNA_ID": canonical["miRNA_ID"],
            "canonical_seqid": canonical["Seq-ID"],
            "isomiR_miRNA_ID": row["miRNA_ID"],
            "isomiR_seqid": row["Seq-ID"],
            **comparison,
        })
    return pairs


def median(values: list[float]) -> float:
    return float(np.median(np.asarray(values, dtype=float))) if values else math.nan


def write_tsv(path: Path, fieldnames: list[str], rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, delimiter="\t", lineterminator="\n", extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def format_p_value(p_value: float) -> str:
    if math.isnan(p_value):
        return "P = NA"
    if p_value < 0.001:
        return f"P = {p_value:.1e}"
    return f"P = {p_value:.3f}"


def main() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    FIGURE_DIR.mkdir(parents=True, exist_ok=True)
    FINAL_FIGURE_DIR.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(WORKBOOK, read_only=True, data_only=True)
    s2_rows = read_sheet(wb, "S2")
    s6_rows = read_sheet(wb, "S6")
    wb.close()

    conservation_by_seqid = {
        clean(row["Seq-ID"]): conservation_label(row["Conservation"])
        for row in s2_rows
    }

    interactions = {}
    for row in s6_rows:
        seqid = clean(row["miRNA_ID"])
        target = clean(row["target_gene"])
        if not seqid or not target:
            continue
        key = (seqid, target, clean(row.get("binding_site_sequence", "")))
        library_count = to_int(row["library_count"])
        if key not in interactions or library_count > interactions[key]["library_count"]:
            interactions[key] = {
                "miRNA_ID": seqid,
                "target_gene": target,
                "library_count": library_count,
                "Conservation": conservation_by_seqid.get(seqid, "Specific"),
            }

    interaction_rows = list(interactions.values())
    target_sets = defaultdict(set)
    for row in interaction_rows:
        target_sets[row["miRNA_ID"]].add(row["target_gene"])

    support_summary = Counter((support_class(row["library_count"]), row["Conservation"]) for row in interaction_rows)
    target_count_by_mirna = []
    for seqid, targets in sorted(target_sets.items()):
        target_count_by_mirna.append({
            "miRNA_ID": seqid,
            "Conservation": conservation_by_seqid.get(seqid, "Specific"),
            "target_count": len(targets),
        })

    pairs = build_canonical_isomir_pairs(s2_rows)
    overlap_rows = []
    for pair in pairs:
        can_targets = target_sets.get(pair["canonical_seqid"], set())
        iso_targets = target_sets.get(pair["isomiR_seqid"], set())
        if not can_targets or not iso_targets:
            continue
        shared = can_targets & iso_targets
        union = can_targets | iso_targets
        overlap_rows.append({
            **pair,
            "canonical_target_count": len(can_targets),
            "isomiR_target_count": len(iso_targets),
            "shared_target_count": len(shared),
            "canonical_specific_target_count": len(can_targets - iso_targets),
            "isomiR_specific_target_count": len(iso_targets - can_targets),
            "union_target_count": len(union),
            "jaccard": len(shared) / len(union) if union else math.nan,
        })

    write_tsv(
        DATA_DIR / "figure4_interaction_support_summary.tsv",
        ["support_class", "Conservation", "interactions"],
        [
            {"support_class": level, "Conservation": cons, "interactions": support_summary[(level, cons)]}
            for level in SUPPORT_LEVELS
            for cons in CONSERVATION_LEVELS
        ],
    )
    write_tsv(DATA_DIR / "figure4_miRNA_target_counts.tsv", ["miRNA_ID", "Conservation", "target_count"], target_count_by_mirna)
    write_tsv(
        DATA_DIR / "isomiR_canonical_target_overlap_pair_detail.tsv",
        [
            "miRNA_loci", "Family", "canonical_miRNA_ID", "canonical_seqid", "isomiR_miRNA_ID", "isomiR_seqid",
            "canonical_target_count", "isomiR_target_count", "shared_target_count",
            "canonical_specific_target_count", "isomiR_specific_target_count", "union_target_count",
            "jaccard", "seed_changed", "terminal_change_5p", "terminal_change_3p",
        ],
        overlap_rows,
    )

    plt.rcParams.update({
        "font.family": "Arial",
        "font.size": 11,
        "axes.linewidth": 1.1,
        "xtick.major.width": 1.1,
        "ytick.major.width": 1.1,
        "figure.facecolor": "white",
    })
    fig = plt.figure(figsize=(10.8, 9.4))
    grid = fig.add_gridspec(2, 2, hspace=0.46, wspace=0.34)

    # A. Interaction support strength.
    ax_a = fig.add_subplot(grid[0, 0])
    x = np.arange(len(SUPPORT_LEVELS))
    bottom = np.zeros(len(SUPPORT_LEVELS))
    colors = {"Conserved": BLUE, "Specific": ORANGE}
    for cons in CONSERVATION_LEVELS:
        values = np.array([support_summary[(level, cons)] for level in SUPPORT_LEVELS])
        ax_a.bar(x, values, bottom=bottom, width=0.72, color=colors[cons], edgecolor="white", linewidth=0.4, label=cons)
        bottom += values
    for idx, total in enumerate(bottom):
        ax_a.text(idx, total + max(bottom) * 0.025, f"{int(total):,}", ha="center", va="bottom", fontsize=9)
    ax_a.set_xticks(x, SUPPORT_LEVELS)
    ax_a.set_xlabel("Number of degradome libraries")
    ax_a.set_ylabel("Number of interactions")
    ax_a.set_title("Interaction support strength", loc="left", fontsize=13, fontweight="bold")
    ax_a.set_ylim(0, 6000)
    ax_a.legend(frameon=False, loc="upper right")
    ax_a.spines[["top", "right"]].set_visible(False)

    # B. Target count distribution.
    ax_b = fig.add_subplot(grid[0, 1])
    max_target_count_bin = 20
    bins = list(range(1, max_target_count_bin)) + [max_target_count_bin]
    medians_by_conservation = {}
    max_bin_height = 0.0
    bin_x = np.arange(len(bins))
    grouped_width = 0.42
    offsets = {"Conserved": -grouped_width / 2, "Specific": grouped_width / 2}
    for cons in CONSERVATION_LEVELS:
        counts = Counter()
        values_for_median = []
        for row in target_count_by_mirna:
            if row["Conservation"] != cons:
                continue
            count = int(row["target_count"])
            values_for_median.append(count)
            counts[min(count, max_target_count_bin)] += 1
        total_mirnas = sum(counts.values())
        bar_values = [counts[b] / total_mirnas if total_mirnas else 0 for b in bins]
        max_bin_height = max(max_bin_height, max(bar_values) if bar_values else 0)
        ax_b.bar(
            bin_x + offsets[cons],
            bar_values,
            width=grouped_width,
            color=colors[cons],
            alpha=0.90,
            label=cons,
            edgecolor="white",
            linewidth=0.25,
        )
        med = median(values_for_median)
        if not math.isnan(med):
            medians_by_conservation[cons] = med
            ax_b.axvline((max_target_count_bin if med >= max_target_count_bin else med) - 1 + offsets[cons], color=colors[cons], linestyle="--", linewidth=1.0)
    ax_b.set_ylim(0, max_bin_height * 1.08 if max_bin_height else 1)
    for cons, med in medians_by_conservation.items():
        x_med = (max_target_count_bin if med >= max_target_count_bin else med) - 1 + offsets[cons]
        y_label = ax_b.get_ylim()[1] * (0.88 if cons == "Conserved" else 0.78)
        ax_b.text(x_med + 0.28, y_label, f"{int(round(med))}", color=colors[cons], fontsize=9, ha="left", va="center")
    tick_positions = [0, 4, 9, 14, 19]
    tick_labels = ["1", "5", "10", "15", "≥20"]
    ax_b.set_xticks(tick_positions, tick_labels)
    ax_b.set_xlabel("Number of target genes per miRNA")
    ax_b.set_ylabel("Proportion of miRNAs")
    ax_b.set_title("Distribution of target-gene counts", loc="left", fontsize=13, fontweight="bold")
    ax_b.legend(frameon=False, loc="upper right")
    ax_b.spines[["top", "right"]].set_visible(False)

    # C. Pairwise target-set similarity.
    ax_c = fig.add_subplot(grid[1, 0])
    jaccards = [float(row["jaccard"]) for row in overlap_rows if not math.isnan(row["jaccard"])]
    ax_c.hist(jaccards, bins=np.linspace(0, 1, 21), color=BLUE, edgecolor="white", linewidth=0.4, alpha=0.88)
    med_j = median(jaccards)
    mean_j = float(np.mean(jaccards)) if jaccards else math.nan
    if not math.isnan(med_j):
        ax_c.axvline(med_j, color=GRAY, linestyle="--", linewidth=1.1)
        ax_c.text(med_j + 0.03, ax_c.get_ylim()[1] * 0.92, f"Median = {med_j:.2f}", fontsize=9, color="black")
    if not math.isnan(mean_j):
        ax_c.axvline(mean_j, color=GRAY, linestyle=":", linewidth=1.3)
        ax_c.text(mean_j + 0.03, ax_c.get_ylim()[1] * 0.82, f"Mean = {mean_j:.2f}", fontsize=9, color="black")
    ax_c.set_xlim(0, 1)
    ax_c.set_xlabel("Jaccard similarity of target sets")
    ax_c.set_ylabel("Number of canonical-isomiR pairs")
    ax_c.set_title("Pairwise target-set similarity", loc="left", fontsize=13, fontweight="bold")
    ax_c.spines[["top", "right"]].set_visible(False)

    # D. Seed-region effect on target overlap.
    ax_d = fig.add_subplot(grid[1, 1])
    box_data = [
        [float(row["jaccard"]) for row in overlap_rows if row["seed_changed"] == "no"],
        [float(row["jaccard"]) for row in overlap_rows if row["seed_changed"] == "yes"],
    ]
    box_labels = ["seed\nunchanged", "seed\nchanged"]
    rng = np.random.default_rng(20260719)
    violins = ax_d.violinplot(
        box_data,
        positions=[1, 2],
        widths=0.72,
        showmeans=False,
        showmedians=False,
        showextrema=False,
    )
    for body, color in zip(violins["bodies"], [BLUE, ORANGE]):
        body.set_facecolor(color)
        body.set_edgecolor("none")
        body.set_alpha(0.25)
    for idx, (values, color) in enumerate(zip(box_data, [BLUE, ORANGE]), start=1):
        if not values:
            continue
        jitter = rng.normal(loc=idx, scale=0.045, size=len(values))
        ax_d.scatter(
            jitter,
            values,
            s=10,
            facecolor=color,
            edgecolor="white",
            linewidth=0.25,
            alpha=0.35,
            zorder=2,
        )
    parts = ax_d.boxplot(
        box_data,
        tick_labels=box_labels,
        patch_artist=True,
        showfliers=False,
        widths=0.30,
        medianprops={"color": "black", "linewidth": 1.25},
        zorder=3,
    )
    for patch, color in zip(parts["boxes"], [BLUE, ORANGE]):
        patch.set_facecolor("none")
        patch.set_edgecolor(color)
        patch.set_linewidth(1.2)
        patch.set_alpha(1.0)
    for idx, values in enumerate(box_data, start=1):
        med = median(values)
        mean_value = float(np.mean(values)) if values else math.nan
        if not math.isnan(mean_value):
            ax_d.scatter(idx, mean_value, s=34, facecolor="white", edgecolor="black", linewidth=0.7, zorder=4)
        if not math.isnan(med):
            ax_d.text(idx + 0.43, med, f"{med:.2f}\n(n={len(values)})", va="center", ha="center", fontsize=9, linespacing=1.15)
    p_seed = mannwhitneyu(box_data[0], box_data[1], alternative="two-sided").pvalue if all(box_data) else math.nan
    bracket_y = 1.04
    bracket_h = 0.035
    ax_d.plot([1, 1, 2, 2], [bracket_y, bracket_y + bracket_h, bracket_y + bracket_h, bracket_y], color="black", linewidth=1.0)
    ax_d.text(1.5, bracket_y + bracket_h + 0.015, f"Mann-Whitney {format_p_value(p_seed)}", ha="center", va="bottom", fontsize=9)
    ax_d.set_ylim(-0.05, 1.15)
    ax_d.set_ylabel("Jaccard similarity")
    ax_d.set_title("Seed-region effect on target overlap", loc="left", fontsize=13, fontweight="bold")
    ax_d.spines[["top", "right"]].set_visible(False)

    for label, axis in zip("ABCD", [ax_a, ax_b, ax_c, ax_d]):
        bbox = axis.get_position()
        fig.text(bbox.x0 - 0.08, bbox.y1 + 0.02, label, fontsize=18, fontweight="bold", ha="left", va="bottom")

    fig.savefig(OUTPUT, dpi=600, bbox_inches="tight")
    fig.savefig(FINAL_OUTPUT, dpi=600, bbox_inches="tight")
    plt.close(fig)

    print(f"S2_rows={len(s2_rows)}")
    print(f"S6_interactions={len(interaction_rows)}")
    print(f"targeted_miRNAs={len(target_sets)}")
    print(f"canonical_isomiR_target_pairs={len(overlap_rows)}")
    print(OUTPUT)
    print(FINAL_OUTPUT)


if __name__ == "__main__":
    main()
