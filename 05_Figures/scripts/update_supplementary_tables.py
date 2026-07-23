#!/usr/bin/env python3
"""
Synchronise supplementary tables S2, S4, S5 with the latest miRNA annotation
workflow results.

Updates:
  S2 — miRNA_ID, annotation_category, locus_class, variant_status, arm_status,
       Reported_Status, Conservation, Family
  S4 — Annotation, Annotation_category
  S5 — family-level summary (recalculated from updated S2)

Matching: Seq-ID + hairpin coordinate (Chr_H_start_H_end), the most stable
record identifier across annotation re-runs.

Usage:
  cd SoymiR-atlas
  python3 05_Figures/scripts/update_supplementary_tables.py
"""

from pathlib import Path
import shutil

import pandas as pd
from openpyxl import load_workbook

SCRIPT_DIR = Path(__file__).resolve().parent
ATLAS_ROOT = SCRIPT_DIR.parents[1]  # SoymiR-atlas/

WORKFLOW_TSV = ATLAS_ROOT / "02_miRNA_annotation/results/2814_precusor_miRNAs_annotated_precursor_overlap_unique_anchor_workflow_short.tsv"
WORKBOOK = ATLAS_ROOT / "05_Figures/input/source_tables/2026_merged_supplementary_tables.xlsx"

# Columns to update in S2 from workflow
S2_SYNC_COLS = [
    "annotation_category",
    "locus_class",
    "variant_status",
    "arm_status",
    "Reported_Status",
    "Conservation",
    "Family",
]
S2_RENAME_MIRNA = {"Annotation": "miRNA_ID"}  # workflow col → S2 col


def main() -> None:
    if not WORKFLOW_TSV.exists():
        raise FileNotFoundError(f"Workflow TSV not found: {WORKFLOW_TSV}")
    if not WORKBOOK.exists():
        raise FileNotFoundError(f"Workbook not found: {WORKBOOK}")

    # Backup
    backup = WORKBOOK.with_suffix(".xlsx.bak")
    shutil.copy2(WORKBOOK, backup)
    print(f"Backup: {backup}")

    # ── Load workflow ──
    wf = pd.read_csv(WORKFLOW_TSV, sep="\t")
    wf["hairpin_key"] = (
        wf["Chr"].astype(str)
        + "_"
        + wf["H_start"].astype(str)
        + "_"
        + wf["H_end"].astype(str)
    )
    wf["coord_key"] = wf["Seq-ID"].astype(str) + "|" + wf["hairpin_key"]

    # Build coordinate → workflow row lookup
    wf_map = {}
    for _, r in wf.iterrows():
        wf_map[r["coord_key"]] = r

    wb = load_workbook(WORKBOOK)

    # ═══════════ S2 ═══════════
    ws2 = wb["S2"]
    h2 = [cell.value for cell in ws2[2]]
    col_seq = h2.index("Seq-ID")
    col_chr = h2.index("Chr")
    col_hs = h2.index("H_start")
    col_he = h2.index("H_end")
    col_indices = {col: h2.index(col) for col in S2_SYNC_COLS}
    col_indices["miRNA_ID"] = h2.index("miRNA_ID")

    s2_updated = 0
    for row in ws2.iter_rows(min_row=3, values_only=False):
        seq_id = str(row[col_seq].value or "")
        chr_v = str(row[col_chr].value or "")
        hs = str(row[col_hs].value or "")
        he = str(row[col_he].value or "")
        ck = f"{seq_id}|{chr_v}_{hs}_{he}"

        if ck not in wf_map:
            continue
        info = wf_map[ck]
        for col_name in S2_SYNC_COLS:
            row[col_indices[col_name]].value = info[col_name]
        # miRNA_ID from workflow Annotation
        row[col_indices["miRNA_ID"]].value = info["Annotation"]
        s2_updated += 1

    print(f"S2: {s2_updated} / {ws2.max_row - 2} records updated")

    # ═══════════ S4 ═══════════
    ws4 = wb["S4"]
    h4 = [cell.value for cell in ws4[2]]
    col4_seq = h4.index("Seq-ID")
    col4_prec = h4.index("Precusor_Coordinate")
    col4_ann = h4.index("Annotation_category")
    col4_mirna = h4.index("Annotation")

    s4_updated = 0
    for row in ws4.iter_rows(min_row=3, values_only=False):
        seq_id = str(row[col4_seq].value or "")
        prec = str(row[col4_prec].value or "")  # Chr_H_start_H_end
        ck = f"{seq_id}|{prec}"

        if ck not in wf_map:
            continue
        info = wf_map[ck]
        row[col4_ann].value = info["annotation_category"]
        row[col4_mirna].value = info["Annotation"]
        s4_updated += 1

    print(f"S4: {s4_updated} / {ws4.max_row - 2} records updated")

    # ═══════════ S5 — recalculate from updated S2 ═══════════
    # Read updated S2 into DataFrame
    s2_data = []
    for row in ws2.iter_rows(min_row=3, values_only=True):
        if row[0]:
            s2_data.append(dict(zip(h2, row)))
    df = pd.DataFrame(s2_data)

    families = (
        df.groupby("Family")
        .agg(
            reference_matched=(
                "annotation_category",
                lambda x: (x == "reference_matched").sum(),
            ),
            nonref_anchor=(
                "annotation_category",
                lambda x: (
                    (df.loc[x.index, "variant_status"] == "anchor")
                    & (x != "reference_matched")
                ).sum(),
            ),
            variant=("variant_status", lambda x: (x == "variant").sum()),
        )
        .reset_index()
    )
    families["Total"] = (
        families["reference_matched"]
        + families["nonref_anchor"]
        + families["variant"]
    )
    families["Variant_percentage"] = (
        families["variant"] / families["Total"] * 100
    ).round(2)

    ws5 = wb["S5"]
    # Clear
    for row in ws5.iter_rows(min_row=3, max_row=ws5.max_row):
        for cell in row:
            cell.value = None
    # Write
    for i, (_, fam) in enumerate(families.iterrows()):
        row_idx = 3 + i
        ws5.cell(row=row_idx, column=1, value=fam["Family"])
        ws5.cell(row=row_idx, column=2, value=int(fam["reference_matched"]))
        ws5.cell(row=row_idx, column=3, value=int(fam["nonref_anchor"]))
        ws5.cell(row=row_idx, column=4, value=int(fam["variant"]))
        ws5.cell(row=row_idx, column=5, value=int(fam["Total"]))
        ws5.cell(row=row_idx, column=6, value=float(fam["Variant_percentage"]))

    print(f"S5: {len(families)} families  |  "
          f"ref={int(families['reference_matched'].sum())}  "
          f"nonref={int(families['nonref_anchor'].sum())}  "
          f"variant={int(families['variant'].sum())}  "
          f"total={int(families['Total'].sum())}")

    # ── Save ──
    wb.save(WORKBOOK)

    # Verify
    s2_fams = {r[col_indices["Family"]] for r in ws2.iter_rows(min_row=3, values_only=True) if r[0]}
    wf_fams = set(wf["Family"].unique())
    assert s2_fams == wf_fams, f"Family mismatch: S2={len(s2_fams)}, WF={len(wf_fams)}"

    print(f"\n✓ Saved: {WORKBOOK}")
    print(f"  S2 families: {len(s2_fams)}  (matches workflow)")


if __name__ == "__main__":
    main()
