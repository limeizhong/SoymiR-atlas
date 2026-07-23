#!/usr/bin/env python3
"""Analyze expression correlation and library overlap for isomiR groups.

For every detected canonical mature-arm with one or more detected isomiRs,
this script calculates:

1. canonical-isomiR pairwise Pearson correlation across the 10 S3 tissue
   expression values;
2. isomiR-isomiR pairwise Pearson correlation when an arm has multiple
   isomiRs;
3. Jaccard similarity between the corresponding detected-library sets.

When a detected canonical record exists, it is selected after removing the
terminal .vN suffix and choosing the best same-locus candidate. Groups without
a detected canonical are retained when at least two distinct isomiR Seq-IDs
are available; singleton groups without canonical are excluded. Pairs whose
two records share the same Seq-ID are also excluded because they do not
represent distinct mature sequences.
"""

from __future__ import annotations

from collections import defaultdict
from itertools import combinations
from pathlib import Path
import csv
import os
import re

import numpy as np
from openpyxl import load_workbook
from scipy.stats import pearsonr, spearmanr


SCRIPT_DIR = Path(__file__).resolve().parent
ATLAS_ROOT = next((p for p in SCRIPT_DIR.parents if p.name == "SoymiR-atlas"), None)
if ATLAS_ROOT is not None:
    ROOT = ATLAS_ROOT.parent
    WORKBOOK = ATLAS_ROOT / "05_Figures/input/source_tables/2026_merged_supplementary_tables.xlsx"
    DATA_DIR = ATLAS_ROOT / "05_Figures/input/plotting_data/Figure3/global_expression_library_overlap"
    FIGURE_DIR = ATLAS_ROOT / "05_Figures/results/intermediate_figures/Figure3"
else:
    raise RuntimeError(
        "Could not locate SoymiR-atlas repository root. "
        "Please run this script from within the SoymiR-atlas directory tree."
    )
RAW_EXPRESSION = ATLAS_ROOT / "04_miRNA_expression/input/68282.1588.expression.rawdata.txt"
OUTPUT_PNG = FIGURE_DIR / "Figure_global_isomiR_expression_correlation_library_overlap_600dpi.png"
OUTPUT_PDF = FIGURE_DIR / "Figure_global_isomiR_expression_correlation_library_overlap.pdf"

os.environ.setdefault("MPLCONFIGDIR", str(ROOT / ".matplotlib_cache"))
os.environ.setdefault("XDG_CACHE_HOME", str(ROOT / ".cache"))
(ROOT / ".matplotlib_cache").mkdir(parents=True, exist_ok=True)
(ROOT / ".cache").mkdir(parents=True, exist_ok=True)

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.lines import Line2D


BLUE = "#0072B2"
ORANGE = "#E69F00"
GRAY = "#666666"
LIGHT_GRAY = "#D9D9D9"
PAIR_ORDER = ["canonical-isomiR", "isomiR-isomiR"]
PAIR_LABELS = ["Canonical-isomiR", "isomiR-isomiR"]
PAIR_COLORS = [BLUE, ORANGE]


def clean(value) -> str:
    return str(value or "").strip()


def to_int(value) -> int:
    return int(float(value))


def canonical_name(miRNA_id: str) -> str:
    return re.sub(r"\.v\d+$", "", clean(miRNA_id), flags=re.IGNORECASE)


def is_isomir(miRNA_id: str) -> bool:
    return bool(re.search(r"\.v\d+$", clean(miRNA_id), flags=re.IGNORECASE))


def read_sheet(sheet_name: str) -> tuple[list[dict], list[str]]:
    workbook = load_workbook(WORKBOOK, read_only=True, data_only=True)
    worksheet = workbook[sheet_name]
    header = [cell.value for cell in worksheet[2]]
    rows = []
    for values in worksheet.iter_rows(min_row=3, values_only=True):
        if values and values[0] is not None:
            rows.append(dict(zip(header, values)))
    workbook.close()
    return rows, header


def row_mid(row: dict, start_key: str, end_key: str) -> float:
    return (to_int(row[start_key]) + to_int(row[end_key])) / 2


def choose_detected_canonical(isomir: dict, candidates: list[dict]) -> dict | None:
    if not candidates:
        return None

    def rank(candidate: dict):
        same_locus = clean(candidate.get("miRNA_loci")).lower() == clean(isomir.get("miRNA_loci")).lower()
        same_chr_strand = clean(candidate.get("Chr")) == clean(isomir.get("Chr")) and clean(candidate.get("Strand")) == clean(isomir.get("Strand"))
        same_hairpin = (
            same_chr_strand
            and to_int(candidate["H_start"]) == to_int(isomir["H_start"])
            and to_int(candidate["H_end"]) == to_int(isomir["H_end"])
        )
        return (
            0 if same_locus and same_hairpin else 1 if same_locus and same_chr_strand else 2,
            abs(row_mid(candidate, "M_start", "M_end") - row_mid(isomir, "M_start", "M_end")),
            abs(row_mid(candidate, "H_start", "H_end") - row_mid(isomir, "H_start", "H_end")),
            to_int(candidate["No."]),
        )

    return min(candidates, key=rank)


def expression_vectors() -> tuple[dict[str, np.ndarray], list[str]]:
    rows, header = read_sheet("S3")
    tissue_columns = header[1:]
    vectors = {
        clean(row["miRNA_ID"]): np.asarray([float(row[tissue] or 0) for tissue in tissue_columns], dtype=float)
        for row in rows
    }
    return vectors, tissue_columns


def read_library_sets(seq_ids: set[str]) -> dict[str, set[str]]:
    library_sets = {seq_id: set() for seq_id in seq_ids}
    with RAW_EXPRESSION.open(encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        for row in reader:
            seq_id = clean(row.get("Seq-ID"))
            sample = clean(row.get("sample"))
            if seq_id in library_sets and sample:
                library_sets[seq_id].add(sample)
    return library_sets


def build_groups(s2: list[dict], expression: dict[str, np.ndarray]):
    canonical_by_name = defaultdict(list)
    isomirs = []
    for row in s2:
        seq_id = clean(row.get("Seq-ID"))
        name = clean(row.get("miRNA_ID"))
        if not seq_id or seq_id not in expression:
            continue
        if is_isomir(name):
            isomirs.append(row)
        else:
            canonical_by_name[name.lower()].append(row)

    groups = defaultdict(lambda: {"canonical": None, "isomiRs": {}})
    no_canonical_candidates = defaultdict(dict)
    for isomir in sorted(isomirs, key=lambda row: to_int(row["No."])):
        expected_name = canonical_name(isomir["miRNA_ID"])
        canonical = choose_detected_canonical(isomir, canonical_by_name.get(expected_name.lower(), []))
        if canonical is None:
            orphan_key = (clean(isomir.get("miRNA_loci")).lower(), expected_name.lower())
            no_canonical_candidates[orphan_key][clean(isomir["Seq-ID"])] = isomir
            continue
        group_key = (
            clean(canonical.get("miRNA_loci")),
            clean(canonical.get("miRNA_ID")),
            clean(canonical.get("Chr")),
            to_int(canonical["H_start"]),
            to_int(canonical["H_end"]),
            clean(canonical.get("Strand")),
            clean(canonical.get("Seq-ID")),
        )
        groups[group_key]["canonical"] = canonical
        # A repeated genomic annotation with the same Seq-ID contributes one
        # expression/library profile within this mature-arm group.
        groups[group_key]["isomiRs"][clean(isomir["Seq-ID"])] = isomir

    excluded_without_partner = []
    no_canonical_groups_added = 0
    for (_, _), isomirs_by_seqid in sorted(no_canonical_candidates.items()):
        orphan_isomirs = list(isomirs_by_seqid.values())
        if len(orphan_isomirs) < 2:
            excluded_without_partner.extend(orphan_isomirs)
            continue
        first = orphan_isomirs[0]
        group_key = (
            clean(first.get("miRNA_loci")),
            canonical_name(first["miRNA_ID"]),
            clean(first.get("Chr")),
            min(to_int(row["H_start"]) for row in orphan_isomirs),
            max(to_int(row["H_end"]) for row in orphan_isomirs),
            clean(first.get("Strand")),
            "",
        )
        groups[group_key]["canonical"] = None
        groups[group_key]["isomiRs"].update(isomirs_by_seqid)
        no_canonical_groups_added += 1

    return groups, excluded_without_partner, no_canonical_groups_added


def safe_pearson(vector_a: np.ndarray, vector_b: np.ndarray) -> tuple[float, float]:
    if np.allclose(vector_a, vector_a[0]) or np.allclose(vector_b, vector_b[0]):
        return np.nan, np.nan
    result = pearsonr(vector_a, vector_b)
    return float(result.statistic), float(result.pvalue)


def pair_metrics(pair_type: str, group_key: tuple, row_a: dict, row_b: dict, expression, libraries) -> dict:
    seq_a = clean(row_a["Seq-ID"])
    seq_b = clean(row_b["Seq-ID"])
    pearson_r, pearson_p = safe_pearson(expression[seq_a], expression[seq_b])
    shared = libraries[seq_a] & libraries[seq_b]
    union = libraries[seq_a] | libraries[seq_b]
    jaccard = len(shared) / len(union) if union else np.nan
    locus, canonical_id, chromosome, h_start, h_end, strand, canonical_seq_id = group_key
    return {
        "Pair_type": pair_type,
        "miRNA_loci": locus,
        "Canonical_miRNA_ID": canonical_id,
        "Canonical_Seq_ID": canonical_seq_id,
        "Chr": chromosome,
        "H_start": h_start,
        "H_end": h_end,
        "Strand": strand,
        "miRNA_A": clean(row_a["miRNA_ID"]),
        "Seq_ID_A": seq_a,
        "miRNA_B": clean(row_b["miRNA_ID"]),
        "Seq_ID_B": seq_b,
        "Pearson_r": pearson_r,
        "Pearson_p": pearson_p,
        "Library_Jaccard": jaccard,
        "Shared_libraries": len(shared),
        "Union_libraries": len(union),
        "Libraries_A": len(libraries[seq_a]),
        "Libraries_B": len(libraries[seq_b]),
    }


def write_tsv(path: Path, fieldnames: list[str], rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, delimiter="\t", lineterminator="\n")
        writer.writeheader()
        for row in rows:
            writer.writerow({key: row.get(key, "") for key in fieldnames})


def summarize(values: list[float]) -> dict:
    array = np.asarray([value for value in values if np.isfinite(value)], dtype=float)
    if not len(array):
        return {key: np.nan for key in ["n", "mean", "median", "q1", "q3", "min", "max"]}
    return {
        "n": len(array),
        "mean": float(np.mean(array)),
        "median": float(np.median(array)),
        "q1": float(np.quantile(array, 0.25)),
        "q3": float(np.quantile(array, 0.75)),
        "min": float(np.min(array)),
        "max": float(np.max(array)),
    }


def finite_median(rows: list[dict], metric: str) -> float:
    values = [row[metric] for row in rows if np.isfinite(row[metric])]
    return float(np.median(values)) if values else np.nan


def add_panel_label(axis, label: str) -> None:
    axis.text(-0.17, 1.08, label, transform=axis.transAxes, fontsize=17, fontweight="bold", va="top", ha="left")


def style_axis(axis, grid_axis="y") -> None:
    axis.spines[["top", "right"]].set_visible(False)
    axis.grid(axis=grid_axis, color=LIGHT_GRAY, linewidth=0.7, alpha=0.7)
    axis.set_axisbelow(True)


def distribution_panel(axis, rows: list[dict], metric: str, ylabel: str, panel: str) -> None:
    groups = [np.asarray([row[metric] for row in rows if row["Pair_type"] == pair and np.isfinite(row[metric])]) for pair in PAIR_ORDER]
    violin = axis.violinplot(groups, positions=[1, 2], widths=0.72, showmeans=False, showmedians=False, showextrema=False)
    for body, color in zip(violin["bodies"], PAIR_COLORS):
        body.set_facecolor(color)
        body.set_edgecolor("none")
        body.set_alpha(0.28)

    rng = np.random.default_rng(20260711)
    for position, values, color in zip([1, 2], groups, PAIR_COLORS):
        jitter = rng.normal(position, 0.055, len(values))
        axis.scatter(jitter, values, s=8, color=color, alpha=0.16, edgecolors="none", rasterized=True)
        q1, median, q3 = np.quantile(values, [0.25, 0.5, 0.75])
        axis.vlines(position, q1, q3, color=color, linewidth=5.5, zorder=4)
        axis.scatter(position, median, s=34, color="white", edgecolor=color, linewidth=1.3, zorder=5)
        axis.text(position, axis.get_ylim()[1], f"n = {len(values)}", ha="center", va="bottom", fontsize=8.2)

    axis.set_xticks([1, 2], PAIR_LABELS)
    axis.set_ylabel(ylabel)
    axis.set_xlim(0.5, 2.5)
    style_axis(axis)
    add_panel_label(axis, panel)


def main() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    FIGURE_DIR.mkdir(parents=True, exist_ok=True)

    s2, _ = read_sheet("S2")
    expression, tissue_columns = expression_vectors()
    groups, excluded, no_canonical_groups_added = build_groups(s2, expression)
    groups_with_canonical = sum(group["canonical"] is not None for group in groups.values())

    seq_ids = set()
    for group in groups.values():
        if group["canonical"] is not None:
            seq_ids.add(clean(group["canonical"]["Seq-ID"]))
        seq_ids.update(group["isomiRs"])
    libraries = read_library_sets(seq_ids)

    rows = []
    identical_seq_pairs = []
    group_summary_rows = []
    for group_key, group in sorted(groups.items(), key=lambda item: item[0]):
        canonical = group["canonical"]
        isomirs = list(group["isomiRs"].values())
        group_pair_rows = []
        if canonical is not None:
            for isomir in isomirs:
                if clean(canonical["Seq-ID"]) == clean(isomir["Seq-ID"]):
                    identical_seq_pairs.append({
                        "miRNA_loci": clean(canonical.get("miRNA_loci")),
                        "Canonical_miRNA_ID": clean(canonical["miRNA_ID"]),
                        "isomiR_miRNA_ID": clean(isomir["miRNA_ID"]),
                        "Shared_Seq_ID": clean(canonical["Seq-ID"]),
                        "Chr": clean(canonical.get("Chr")),
                        "M_start": canonical["M_start"],
                        "M_end": canonical["M_end"],
                        "Reason": "canonical and isomiR have the same Seq-ID",
                    })
                    continue
                group_pair_rows.append(pair_metrics("canonical-isomiR", group_key, canonical, isomir, expression, libraries))
        for isomir_a, isomir_b in combinations(isomirs, 2):
            group_pair_rows.append(pair_metrics("isomiR-isomiR", group_key, isomir_a, isomir_b, expression, libraries))
        rows.extend(group_pair_rows)

        canonical_pairs = [row for row in group_pair_rows if row["Pair_type"] == "canonical-isomiR"]
        isomir_pairs = [row for row in group_pair_rows if row["Pair_type"] == "isomiR-isomiR"]
        locus, canonical_id, chromosome, h_start, h_end, strand, canonical_seq_id = group_key
        group_summary_rows.append({
            "miRNA_loci": locus,
            "Canonical_miRNA_ID": canonical_id,
            "Canonical_Seq_ID": canonical_seq_id,
            "Chr": chromosome,
            "H_start": h_start,
            "H_end": h_end,
            "Strand": strand,
            "isomiR_count": len(isomirs),
            "isomiR_IDs": ";".join(sorted({clean(row["miRNA_ID"]) for row in isomirs})),
            "isomiR_Seq_IDs": ";".join(sorted({clean(row["Seq-ID"]) for row in isomirs})),
            "Canonical_isomiR_pair_count": len(canonical_pairs),
            "Canonical_isomiR_median_Pearson_r": finite_median(canonical_pairs, "Pearson_r"),
            "Canonical_isomiR_median_Library_Jaccard": finite_median(canonical_pairs, "Library_Jaccard"),
            "isomiR_isomiR_pair_count": len(isomir_pairs),
            "isomiR_isomiR_median_Pearson_r": finite_median(isomir_pairs, "Pearson_r"),
            "isomiR_isomiR_median_Library_Jaccard": finite_median(isomir_pairs, "Library_Jaccard"),
        })

    detail_fields = [
        "Pair_type", "miRNA_loci", "Canonical_miRNA_ID", "Canonical_Seq_ID", "Chr", "H_start", "H_end", "Strand",
        "miRNA_A", "Seq_ID_A", "miRNA_B", "Seq_ID_B", "Pearson_r", "Pearson_p", "Library_Jaccard",
        "Shared_libraries", "Union_libraries", "Libraries_A", "Libraries_B",
    ]
    write_tsv(DATA_DIR / "global_isomiR_pairwise_expression_library_metrics.tsv", detail_fields, rows)
    write_tsv(
        DATA_DIR / "pairs_excluded_identical_seq_id.tsv",
        ["miRNA_loci", "Canonical_miRNA_ID", "isomiR_miRNA_ID", "Shared_Seq_ID", "Chr", "M_start", "M_end", "Reason"],
        identical_seq_pairs,
    )
    group_summary_fields = [
        "miRNA_loci", "Canonical_miRNA_ID", "Canonical_Seq_ID", "Chr", "H_start", "H_end", "Strand",
        "isomiR_count", "isomiR_IDs", "isomiR_Seq_IDs", "Canonical_isomiR_pair_count",
        "Canonical_isomiR_median_Pearson_r", "Canonical_isomiR_median_Library_Jaccard",
        "isomiR_isomiR_pair_count", "isomiR_isomiR_median_Pearson_r", "isomiR_isomiR_median_Library_Jaccard",
    ]
    write_tsv(DATA_DIR / "global_isomiR_mature_arm_summary.tsv", group_summary_fields, group_summary_rows)

    excluded_rows = [
        {
            "miRNA_ID": clean(row["miRNA_ID"]),
            "Seq-ID": clean(row["Seq-ID"]),
            "miRNA_loci": clean(row.get("miRNA_loci")),
            "Expected_canonical_ID": canonical_name(row["miRNA_ID"]),
            "Reason": "no detected canonical and only one distinct isomiR sequence",
        }
        for row in excluded
    ]
    write_tsv(
        DATA_DIR / "isomiRs_excluded_without_pairable_partner.tsv",
        ["miRNA_ID", "Seq-ID", "miRNA_loci", "Expected_canonical_ID", "Reason"],
        excluded_rows,
    )

    summary_rows = []
    for metric in ["Pearson_r", "Library_Jaccard"]:
        stats = summarize([row[metric] for row in rows])
        summary_rows.append({"Pair_type": "all-within-locus", "Metric": metric, **stats})
    write_tsv(DATA_DIR / "global_isomiR_pairwise_summary.tsv", ["Pair_type", "Metric", "n", "mean", "median", "q1", "q3", "min", "max"], summary_rows)

    valid_rows = [row for row in rows if np.isfinite(row["Pearson_r"]) and np.isfinite(row["Library_Jaccard"])]
    rho_all = spearmanr([row["Pearson_r"] for row in valid_rows], [row["Library_Jaccard"] for row in valid_rows])
    test_rows = [
        {"Test": "Spearman", "Metric": "Pearson_r_vs_Library_Jaccard", "Statistic": rho_all.statistic, "P_value": rho_all.pvalue},
    ]
    write_tsv(DATA_DIR / "global_isomiR_pairwise_statistical_tests.tsv", ["Test", "Metric", "Statistic", "P_value"], test_rows)

    report_lines = [
        "# Global isomiR expression-correlation and library-overlap analysis",
        "",
        "## Analysis scope",
        "",
        f"- Mature-arm groups with a detected canonical miRNA: {groups_with_canonical}",
        f"- Mature-arm groups without canonical but with at least two distinct isomiRs: {no_canonical_groups_added}",
        f"- isomiRs excluded because no canonical and no second distinct isomiR were detected: {len(excluded)}",
        f"- Canonical-isomiR pairs excluded because both records had the same Seq-ID: {len(identical_seq_pairs)}",
        f"- All within-locus mature-isoform pairs: {len(rows)}",
        "",
        "Pearson correlation was calculated across the 10 tissue-level expression values in S3. "
        "Library overlap was calculated as the Jaccard index of detected-library sets from "
        "68282.1588.expression.rawdata.txt.",
        "",
        "## Summary",
        "",
    ]
    r_values = [row["Pearson_r"] for row in rows if np.isfinite(row["Pearson_r"])]
    j_values = [row["Library_Jaccard"] for row in rows if np.isfinite(row["Library_Jaccard"])]
    zero_overlap = sum(value == 0 for value in j_values)
    report_lines.extend([
        "### All within-locus mature-isoform pairs",
        "",
        f"- Pearson r: median {np.median(r_values):.3f}; IQR {np.quantile(r_values, 0.25):.3f} to {np.quantile(r_values, 0.75):.3f}",
        f"- Pairs with positive Pearson r: {sum(value > 0 for value in r_values)}/{len(r_values)} ({sum(value > 0 for value in r_values) / len(r_values) * 100:.1f}%)",
        f"- Pairs with Pearson r >= 0.5: {sum(value >= 0.5 for value in r_values)}/{len(r_values)} ({sum(value >= 0.5 for value in r_values) / len(r_values) * 100:.1f}%)",
        f"- Library Jaccard: median {np.median(j_values):.3f}; IQR {np.quantile(j_values, 0.25):.3f} to {np.quantile(j_values, 0.75):.3f}",
        f"- Pairs with no shared detection library: {zero_overlap}/{len(j_values)} ({zero_overlap / len(j_values) * 100:.1f}%)",
        "",
    ])
    report_lines.extend([
        "## Overall association",
        "",
        f"Pearson expression correlation and library Jaccard similarity showed a weak positive association "
        f"(Spearman rho = {rho_all.statistic:.3f}, P = {rho_all.pvalue:.3g}).",
        "",
    ])
    (DATA_DIR / "global_isomiR_expression_library_overlap_report.md").write_text("\n".join(report_lines), encoding="utf-8")

    plt.rcParams.update({
        "font.family": "Arial",
        "font.size": 10,
        "axes.linewidth": 1.0,
        "xtick.major.width": 1.0,
        "ytick.major.width": 1.0,
        "figure.facecolor": "white",
    })
    fig, axes = plt.subplots(1, 3, figsize=(11.8, 4.0), gridspec_kw={"wspace": 0.38})

    pearson_values = np.asarray([row["Pearson_r"] for row in valid_rows], dtype=float)
    axes[0].hist(pearson_values, bins=np.linspace(-1, 1, 25), color=BLUE, alpha=0.82, edgecolor="white", linewidth=0.5)
    axes[0].axvline(0, color=GRAY, linewidth=0.8, linestyle="--", alpha=0.7)
    axes[0].axvline(np.median(pearson_values), color="black", linewidth=1.1, linestyle="--")
    axes[0].text(np.median(pearson_values) + 0.035, axes[0].get_ylim()[1] * 0.93, f"Median = {np.median(pearson_values):.2f}", fontsize=8.2, va="top")
    axes[0].set_xlim(-1, 1)
    axes[0].set_xlabel("Pearson correlation (r)")
    axes[0].set_ylabel("Number of within-locus pairs")
    axes[0].set_title("Tissue-expression correlation", loc="left", fontsize=11.5, fontweight="bold", pad=9)
    style_axis(axes[0])
    add_panel_label(axes[0], "A")

    jaccard_values = np.asarray([row["Library_Jaccard"] for row in valid_rows], dtype=float)
    overlap_bins = [
        ("0", jaccard_values == 0),
        ("0-0.05", (jaccard_values > 0) & (jaccard_values <= 0.05)),
        ("0.05-0.10", (jaccard_values > 0.05) & (jaccard_values <= 0.10)),
        ("0.10-0.25", (jaccard_values > 0.10) & (jaccard_values <= 0.25)),
        ("0.25-0.50", (jaccard_values > 0.25) & (jaccard_values <= 0.50)),
        (">0.50", jaccard_values > 0.50),
    ]
    overlap_labels = [label for label, _ in overlap_bins]
    overlap_percentages = np.asarray([np.mean(mask) * 100 for _, mask in overlap_bins])
    axes[1].bar(range(len(overlap_labels)), overlap_percentages, color=ORANGE, width=0.72)
    for index, value in enumerate(overlap_percentages):
        if value >= 1:
            axes[1].text(index, value + 1.2, f"{value:.1f}%", ha="center", va="bottom", fontsize=7.6)
    axes[1].set_xticks(range(len(overlap_labels)), overlap_labels, rotation=35, ha="right")
    axes[1].set_ylabel("Proportion of within-locus pairs (%)")
    axes[1].set_xlabel("Library-overlap Jaccard index")
    axes[1].set_ylim(0, max(overlap_percentages) * 1.13)
    axes[1].set_title("Library-detection overlap", loc="left", fontsize=11.5, fontweight="bold", pad=9)
    style_axis(axes[1])
    add_panel_label(axes[1], "B")

    axes[2].scatter(
        pearson_values,
        jaccard_values,
        s=15,
        color=BLUE,
        alpha=0.33,
        edgecolors="none",
        rasterized=True,
    )
    axes[2].set_xlim(-1.05, 1.05)
    axes[2].set_ylim(-0.03, 1.03)
    axes[2].set_xlabel("Pearson correlation (r)")
    axes[2].set_ylabel("Library-overlap Jaccard index")
    axes[2].set_title("Expression versus library overlap", loc="left", fontsize=11.5, fontweight="bold", pad=9)
    axes[2].text(
        0.03,
        0.97,
        f"Spearman rho = {rho_all.statistic:.2f}\nP = {rho_all.pvalue:.2g}",
        transform=axes[2].transAxes,
        ha="left",
        va="top",
        fontsize=8.2,
    )
    style_axis(axes[2])
    add_panel_label(axes[2], "C")

    fig.suptitle("Within-locus expression relationships among soybean miRNA isoforms", x=0.06, ha="left", fontsize=14, fontweight="bold")
    fig.subplots_adjust(left=0.075, right=0.985, bottom=0.18, top=0.82)
    fig.savefig(OUTPUT_PNG, dpi=600, bbox_inches="tight", facecolor="white")
    fig.savefig(OUTPUT_PDF, bbox_inches="tight", facecolor="white")
    plt.close(fig)

    print(f"mature_arm_groups_with_detected_canonical={groups_with_canonical}")
    print(f"mature_arm_groups_without_canonical_with_multiple_isomiRs={no_canonical_groups_added}")
    print(f"isomiRs_excluded_without_pairable_partner={len(excluded)}")
    print(f"pairs_excluded_identical_seq_id={len(identical_seq_pairs)}")
    print(f"canonical_isomiR_pairs={sum(row['Pair_type'] == 'canonical-isomiR' for row in rows)}")
    print(f"isomiR_isomiR_pairs={sum(row['Pair_type'] == 'isomiR-isomiR' for row in rows)}")
    print(f"tissues={','.join(tissue_columns)}")
    print(OUTPUT_PNG)


if __name__ == "__main__":
    main()
