#!/usr/bin/env python3
from __future__ import annotations

from collections import Counter, defaultdict
from pathlib import Path
import csv
import os
import re

import numpy as np
from openpyxl import load_workbook
from scipy.stats import mannwhitneyu, pearsonr

SCRIPT_DIR = Path(__file__).resolve().parent
ATLAS_ROOT = next((p for p in SCRIPT_DIR.parents if p.name == "SoymiR-atlas"), None)

if ATLAS_ROOT is not None:
    ROOT = ATLAS_ROOT.parent
    WORKBOOK = ATLAS_ROOT / "05_Figures/input/source_tables/2026_merged_supplementary_tables.xlsx"
    DATA_DIR = ATLAS_ROOT / "05_Figures/input/plotting_data/Figure3"
    FIGURE_DIR = ATLAS_ROOT / "05_Figures/results/intermediate_figures/Figure3"
    FINAL_FIGURE_DIR = ATLAS_ROOT / "05_Figures/results/final_figures"
    RAW_EXPRESSION_FILE = DATA_DIR / "68282.1588.expression.rawdata.txt"
    CANONICAL_FASTA = DATA_DIR / "gma_mature_wo_U.fa"
    MIRBASE_FASTA = DATA_DIR / "gma-mature.fa"
    PMIREN_FASTA = DATA_DIR / "pmiren_gmax_mature.fa"
else:
    raise RuntimeError(
        "Could not locate SoymiR-atlas repository root. "
        "Please run this script from within the SoymiR-atlas directory tree."
    )

OUTPUT = FIGURE_DIR / "Figure_isomiR_composition_detection_rich_families_global_panel_E_600dpi.png"
FINAL_OUTPUT = FINAL_FIGURE_DIR / "Figure_3.png"
GLOBAL_PAIR_FILE = DATA_DIR / "global_expression_library_overlap/global_isomiR_pairwise_expression_library_metrics.tsv"

os.environ.setdefault("MPLCONFIGDIR", str(ROOT / ".matplotlib_cache"))
os.environ.setdefault("XDG_CACHE_HOME", str(ROOT / ".cache"))
(ROOT / ".matplotlib_cache").mkdir(parents=True, exist_ok=True)
(ROOT / ".cache").mkdir(parents=True, exist_ok=True)

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.colors import LinearSegmentedColormap

BLUE = "#4C78A8"
ORANGE = "#E17C05"
GREEN = "#009E73"
# Light yellow-orange corresponding approximately to the C-panel heatmap
# color around the value 80; used only for the isomiR segment in panel B.
ISOMIR_BAR = "#FEC44F"
GRAY = "#777777"
EXPRESSION_CMAP = LinearSegmentedColormap.from_list("soft_expression", ["#FFFFFF", "#FFF7BC", "#FDAE6B", "#F03B20"])
CORRELATION_CMAP = LinearSegmentedColormap.from_list("blue_white_orange", ["#9BB9D5", "#F7F7F7", "#F0B15A"])
C_HEATMAP_CMAP = LinearSegmentedColormap.from_list("soft_yellow_orange_red", ["#FFFFFF", "#FFF7BC", "#FEC44F", "#F03B20"])
PANEL_E_TISSUE_ORDER = ["leaves", "seed", "flower", "root", "stem", "pod", "cotyledon", "shoot", "nodules", "mixed"]
PANEL_E_TISSUE_LABELS = ["Leaves", "Seed", "Flower", "Root", "Stem", "Pod", "Cotyledon", "Shoot", "Nodules", "Mixed"]


def is_isomir(row: dict) -> bool:
    variant_status = str(row.get("variant_status", ""))
    name = str(row.get("miRNA_ID", ""))
    return variant_status == "variant" or bool(re.search(r"\.v\d+$", name))


def read_workbook():
    wb = load_workbook(WORKBOOK, read_only=True, data_only=True)

    ws2 = wb["S2"]
    h2 = [cell.value for cell in ws2[2]]
    s2 = []
    for row in ws2.iter_rows(min_row=3, values_only=True):
        if not row[0]:
            continue
        rec = dict(zip(h2, row))
        rec["Group"] = "isomiR" if is_isomir(rec) else "canonical"
        rec["Library_detection_count"] = library_count_from_seq_id(rec["Seq-ID"])
        s2.append(rec)

    ws3 = wb["S3"]
    h3 = [cell.value for cell in ws3[2]]
    expression = {}
    for row in ws3.iter_rows(min_row=3, values_only=True):
        if row[0]:
            expression[row[0]] = np.asarray([float(v or 0) for v in row[1:]], dtype=float)
    return s2, expression, h3[1:]


def library_count_from_seq_id(seq_id: str) -> int:
    match = re.search(r"_R(\d+)_", str(seq_id))
    return int(match.group(1)) if match else 0


def detection_bin(count: int) -> str:
    if count == 1:
        return "1"
    if count <= 5:
        return "2-5"
    if count <= 10:
        return "6-10"
    if count <= 50:
        return "11-50"
    if count <= 100:
        return "51-100"
    return ">100"


def style_axis(axis, grid_axis="y"):
    axis.spines[["top", "right"]].set_visible(False)
    axis.grid(axis=grid_axis, color="#D9D9D9", linewidth=0.7, alpha=0.7)
    axis.set_axisbelow(True)


def add_panel_label(fig, axis, label):
    bbox = axis.get_position()
    x = 0.02 if bbox.x0 < 0.5 else 0.52
    if label in {"A", "C", "E"}:
        x += 0.011
    fig.text(x, bbox.y1 + 0.012, label, fontsize=18, fontweight="bold", va="bottom", ha="left")


def significance_label(p_value):
    if p_value < 0.001:
        return "***"
    if p_value < 0.01:
        return "**"
    if p_value < 0.05:
        return "*"
    return ""


def write_tsv(path: Path, fieldnames: list[str], rows: list[dict]):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, delimiter="\t", lineterminator="\n", extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def read_global_pair_values() -> tuple[np.ndarray, np.ndarray]:
    pearson = []
    jaccard = []
    with GLOBAL_PAIR_FILE.open(encoding="utf-8", newline="") as handle:
        for row in csv.DictReader(handle, delimiter="\t"):
            r_value = float(row["Pearson_r"])
            j_value = float(row["Library_Jaccard"])
            if np.isfinite(r_value) and np.isfinite(j_value):
                pearson.append(r_value)
                jaccard.append(j_value)
    return np.asarray(pearson), np.asarray(jaccard)


def read_miR166g_v3_v4_point() -> tuple[float, float]:
    with GLOBAL_PAIR_FILE.open(encoding="utf-8", newline="") as handle:
        for row in csv.DictReader(handle, delimiter="\t"):
            names = {row["miRNA_A"], row["miRNA_B"]}
            if names == {"gma-miR166g.v3", "gma-miR166g.v4"}:
                return float(row["Pearson_r"]), float(row["Library_Jaccard"])
    raise ValueError("Could not find gma-miR166g.v3 vs gma-miR166g.v4 in pairwise metrics")


def read_fasta(path: Path) -> dict[str, str]:
    records = {}
    current_id = None
    chunks = []
    with path.open(encoding="utf-8") as handle:
        for line in handle:
            line = line.strip()
            if not line:
                continue
            if line.startswith(">"):
                if current_id is not None:
                    records[current_id] = "".join(chunks).upper().replace("U", "T")
                current_id = line[1:].split()[0]
                chunks = []
            else:
                chunks.append(line)
    if current_id is not None:
        records[current_id] = "".join(chunks).upper().replace("U", "T")
    return records


def canonical_name(miRNA_id: str) -> str:
    return re.sub(r"\.v\d+$", "", str(miRNA_id))


def to_int(value) -> int:
    return int(float(value))


def row_mid(row: dict, start_key: str, end_key: str) -> float:
    return (to_int(row[start_key]) + to_int(row[end_key])) / 2


def pair_table_row(row: dict, group: str, expected_canonical_id: str | None = None) -> dict:
    out = {
        "No.": row["No."],
        "miRNA_ID": row["miRNA_ID"],
        "Chr": row["Chr"],
        "M_start": row["M_start"],
        "M_end": row["M_end"],
        "H_start": row["H_start"],
        "H_end": row["H_end"],
        "Strand": row["Strand"],
        "Seq-ID": row["Seq-ID"],
        "Sequences_Mature": row["Sequences_Mature"],
        "Family": row["Family"],
        "miRNA_loci": row["miRNA_loci"],
        "Hairpin_loci": f"{row['Chr']}:{row['H_start']}-{row['H_end']}:{row['Strand']}",
        "miRBase-mature": "",
        "miRBase-hairpin": "",
        "Report_status": str(row["Reported_Status"]).capitalize(),
        "Conservation": str(row["Conservation"]).capitalize(),
        "Group": group,
    }
    if expected_canonical_id is not None:
        out["Expected_canonical_ID"] = expected_canonical_id
    return out


def choose_detected_canonical(isomir: dict, candidates: list[dict]) -> dict | None:
    if not candidates:
        return None

    def rank(candidate: dict):
        same_locus = candidate["miRNA_loci"] == isomir["miRNA_loci"]
        same_chr_strand = candidate["Chr"] == isomir["Chr"] and candidate["Strand"] == isomir["Strand"]
        same_hairpin = (
            same_chr_strand
            and candidate["H_start"] == isomir["H_start"]
            and candidate["H_end"] == isomir["H_end"]
        )
        return (
            0 if same_locus and same_hairpin else 1 if same_locus and same_chr_strand else 2,
            abs(row_mid(candidate, "M_start", "M_end") - row_mid(isomir, "M_start", "M_end")),
            abs(row_mid(candidate, "H_start", "H_end") - row_mid(isomir, "H_start", "H_end")),
        )

    return min(candidates, key=rank)


def build_canonical_isomir_pair_tables(s2: list[dict]) -> tuple[list[dict], list[dict]]:
    canonical_by_id = defaultdict(list)
    isomirs = []
    for row in s2:
        if row["Group"] == "isomiR":
            isomirs.append(row)
        else:
            canonical_by_id[row["miRNA_ID"]].append(row)

    pair_rows = []
    no_canonical_rows = []
    seen_canonical = set()
    for row in sorted(isomirs, key=lambda r: (str(r["miRNA_loci"]), to_int(r["No."]))):
        expected_id = canonical_name(row["miRNA_ID"])
        canonical = choose_detected_canonical(row, canonical_by_id.get(expected_id, []))
        if canonical is not None:
            key = tuple(canonical[k] for k in ["No.", "miRNA_ID", "Chr", "M_start", "M_end", "Seq-ID"])
            if key not in seen_canonical:
                pair_rows.append(pair_table_row(canonical, "canonical"))
                seen_canonical.add(key)
        else:
            no_canonical_rows.append(pair_table_row(row, "isomiR", expected_id))
        pair_rows.append(pair_table_row(row, "isomiR"))

    pair_rows.sort(key=lambda r: (str(r["miRNA_loci"]), 0 if r["Group"] == "canonical" else 1, to_int(r["No."])))
    no_canonical_rows.sort(key=lambda r: (str(r["miRNA_loci"]), to_int(r["No."])))
    return pair_rows, no_canonical_rows


def build_detection_pair_rows(s2: list[dict]) -> list[dict]:
    canonical_by_id = defaultdict(list)
    for row in s2:
        if row["Group"] == "canonical":
            canonical_by_id[row["miRNA_ID"]].append(row)

    rows = []
    for row in sorted((r for r in s2 if r["Group"] == "isomiR"), key=lambda r: to_int(r["No."])):
        expected_id = canonical_name(row["miRNA_ID"])
        canonical = choose_detected_canonical(row, canonical_by_id.get(expected_id, []))
        rows.append({
            "Canonical_miRNA_ID": canonical["miRNA_ID"] if canonical is not None else expected_id,
            "Canonical_Seq_ID": canonical["Seq-ID"] if canonical is not None else "",
            "Canonical_library_detection_count": library_count_from_seq_id(canonical["Seq-ID"]) if canonical is not None else 0,
            "IsomiR_miRNA_ID": row["miRNA_ID"],
            "IsomiR_Seq_ID": row["Seq-ID"],
            "IsomiR_library_detection_count": library_count_from_seq_id(row["Seq-ID"]),
            "Canonical_detected": "yes" if canonical is not None else "no",
            "Family": row["Family"],
            "miRNA_loci": row["miRNA_loci"],
        })
    return rows


def coordinate_offsets(canonical: dict, isomir: dict) -> tuple[int, int]:
    can_start = to_int(canonical["M_start"])
    can_end = to_int(canonical["M_end"])
    iso_start = to_int(isomir["M_start"])
    iso_end = to_int(isomir["M_end"])
    if canonical["Strand"] == "+":
        return iso_start - can_start, iso_end - can_end
    return can_end - iso_end, can_start - iso_start


def best_ungapped_offsets(canonical_seq: str, isomir_seq: str) -> tuple[int, int, int, int]:
    can = canonical_seq.upper().replace("U", "T")
    iso = isomir_seq.upper().replace("U", "T")
    best = None
    for offset_5p in range(-len(iso), len(can) + 1):
        can_start = max(0, offset_5p)
        iso_start = max(0, -offset_5p)
        overlap = min(len(can) - can_start, len(iso) - iso_start)
        if overlap <= 0:
            continue
        matches = sum(can[can_start + i] == iso[iso_start + i] for i in range(overlap))
        mismatches = overlap - matches
        score = matches - mismatches - 0.05 * abs(offset_5p)
        candidate = (score, matches, -mismatches, overlap, offset_5p)
        if best is None or candidate > best:
            best = candidate
    if best is None:
        return 0, len(iso) - len(can), 0, min(len(can), len(iso))
    _, _matches, neg_mismatches, overlap, offset_5p = best
    offset_3p = offset_5p + len(iso) - len(can)
    return offset_5p, offset_3p, -neg_mismatches, overlap


def compare_sequences(canonical_seq: str, isomir_seq: str, offset_5p=None, offset_3p=None) -> dict:
    can = canonical_seq.upper().replace("U", "T")
    iso = isomir_seq.upper().replace("U", "T")
    if offset_5p is None or offset_3p is None:
        offset_5p, offset_3p, mismatches, overlap = best_ungapped_offsets(can, iso)
    else:
        can_start = max(0, offset_5p)
        iso_start = max(0, -offset_5p)
        overlap = min(len(can) - can_start, len(iso) - iso_start)
        mismatches = sum(can[can_start + i] != iso[iso_start + i] for i in range(overlap)) if overlap > 0 else 0
    return {
        "Canonical_sequence": can,
        "IsomiR_sequence": iso,
        "Canonical_length": len(can),
        "IsomiR_length": len(iso),
        "Length_delta": len(iso) - len(can),
        "Offset_5p": offset_5p,
        "Offset_3p": offset_3p,
        "Terminal_change_5p": -offset_5p,
        "Terminal_change_3p": offset_3p,
        "Overlap_mismatches": mismatches,
        "Overlap_length": overlap,
    }


def build_rich_family_summary(s2):
    families = defaultdict(lambda: {"Reported": 0, "Unreported": 0, "isomiRs": 0})
    for row in s2:
        fam = row["Family"]
        if row["Group"] == "isomiR":
            families[fam]["isomiRs"] += 1
        elif row["Reported_Status"] == "reported":
            families[fam]["Reported"] += 1
        else:
            families[fam]["Unreported"] += 1
    rows = []
    for fam, counts in families.items():
        total = counts["Reported"] + counts["Unreported"] + counts["isomiRs"]
        rows.append({
            "Family": fam,
            "Reported": counts["Reported"],
            "Unreported": counts["Unreported"],
            "isomiRs": counts["isomiRs"],
            "Total": total,
            "isomiR_percentage": f"{counts['isomiRs'] / total * 100:.2f}" if total else "0.00",
        })
    rows.sort(key=lambda r: (-int(r["isomiRs"]), -int(r["Total"]), r["Family"]))
    return rows


def get_fasta_sequence(fasta: dict[str, str], mature_id: str) -> tuple[str, str | None, str]:
    if mature_id in fasta:
        return mature_id, fasta[mature_id], "exact_fasta_id"
    lower_lookup = {key.lower(): key for key in fasta}
    key = lower_lookup.get(mature_id.lower())
    if key:
        return key, fasta[key], "case_insensitive_fasta_id"
    return mature_id, None, "missing_in_fasta"


def terminal_changes(pair_rows: list[dict], no_canonical_rows: list[dict], s2: list[dict]):
    fasta_original = read_fasta(CANONICAL_FASTA)
    fasta_mirbase = read_fasta(MIRBASE_FASTA) if MIRBASE_FASTA.exists() else {}
    fasta_pmiren = read_fasta(PMIREN_FASTA) if PMIREN_FASTA.exists() else {}

    canonical_by_id = defaultdict(list)
    for row in pair_rows:
        if row["Group"] == "canonical":
            canonical_by_id[row["miRNA_ID"]].append(row)

    # Build full S2 canonical lookup: canonical miRNA_ID → Sequences_Mature
    # Also record the source for each isomiR row
    s2_canonical_seq = {}
    isomiR_source = {}
    for row in s2:
        if row["Group"] == "canonical":
            sid = str(row.get("miRNA_ID") or "").strip()
            seq = str(row.get("Sequences_Mature") or "").strip()
            if sid and seq:
                s2_canonical_seq[sid] = seq
        else:
            sid = str(row.get("miRNA_ID") or "").strip()
            src = str(row.get("source") or "").strip().lower()
            if sid:
                isomiR_source[sid] = src

    expected_by_key = {
        (row["miRNA_ID"], row["Seq-ID"]): row["Expected_canonical_ID"]
        for row in no_canonical_rows
    }

    def lookup_fasta(canonical_id: str, source: str) -> tuple[str | None, str]:
        """Search FASTA files in priority order based on annotation source.
        Returns (fasta_id, sequence) or (None, 'missing') if not found.
        Handles PmiREN pipe-delimited headers (e.g. Gma-MIR12209|PmiREN...) and
        MIR/miR naming differences."""
        # Determine FASTA search order based on source
        fasta_list = []
        if source == "mirbase":
            fasta_list = [("gma-mature.fa", fasta_mirbase)]
        elif source == "pmiren":
            fasta_list = [("pmiren_gmax_mature.fa", fasta_pmiren)]
        # Fallback: try all
        fasta_list += [
            ("gma-mature.fa", fasta_mirbase),
            ("pmiren_gmax_mature.fa", fasta_pmiren),
            ("gma_mature_wo_U.fa", fasta_original),
        ]

        def _normalize(name: str) -> str:
            """Normalize miRNA names for fuzzy matching."""
            return name.lower().replace("mir", "mir").replace("-", "")

        def _match_key(fasta_key: str, query: str) -> bool:
            """Check if fasta_key matches the query, considering pipe-delimited
            headers and MIR/miR differences."""
            # Exact match
            if fasta_key == query:
                return True
            # Case-insensitive
            if fasta_key.lower() == query.lower():
                return True
            # First | field (PmiREN short ID), case-insensitive
            short = fasta_key.split("|")[0]
            if short.lower() == query.lower():
                return True
            # MIR → miR normalization (Gma-MIR12209 → Gma-miR12209)
            normalized_short = short.lower().replace("mir", "mir")
            normalized_query = query.lower().replace("mir", "mir")
            if normalized_short == normalized_query:
                return True
            return False

        for _fname, fdict in fasta_list:
            if not fdict:
                continue
            for key in fdict:
                if _match_key(key, canonical_id):
                    return key, fdict[key]
        return None, "missing"

    counts = Counter()
    detail = []
    for row in pair_rows:
        if row["Group"] != "isomiR":
            continue
        expected_id = canonical_name(row["miRNA_ID"])
        canonical = choose_detected_canonical(row, canonical_by_id.get(expected_id, []))
        if canonical is not None:
            offset_5p, offset_3p = coordinate_offsets(canonical, row)
            comparison = compare_sequences(canonical["Sequences_Mature"], row["Sequences_Mature"], offset_5p, offset_3p)
            fasta_id = canonical["miRNA_ID"]
            fasta_match_type = "detected_canonical_sequence"
            canonical_id = canonical["miRNA_ID"]
            canonical_source = "detected_canonical_row"
        else:
            canonical_id = expected_by_key.get((row["miRNA_ID"], row["Seq-ID"]), expected_id)
            # Priority: S2 table → source-specific FASTA → all FASTA
            s2_seq = s2_canonical_seq.get(canonical_id)
            if s2_seq is not None:
                comparison = compare_sequences(s2_seq, row["Sequences_Mature"])
                fasta_id = canonical_id
                fasta_match_type = "canonical_from_S2"
                canonical_source = "canonical_from_S2"
            else:
                src = isomiR_source.get(row["miRNA_ID"], "")
                fasta_id, canonical_seq = lookup_fasta(canonical_id, src)
                if canonical_seq is None:
                    continue
                comparison = compare_sequences(canonical_seq, row["Sequences_Mature"])
                fasta_match_type = "canonical_from_fasta"
                canonical_source = f"canonical_from_fasta_{src}" if src else "canonical_from_fasta"

        overlap = int(comparison["Overlap_length"])
        mismatches = int(comparison["Overlap_mismatches"])
        identity = (overlap - mismatches) / overlap if overlap > 0 else 0
        if overlap < 10 or identity < 0.70:
            continue

        change_5p = comparison["Terminal_change_5p"]
        change_3p = comparison["Terminal_change_3p"]
        counts[(change_5p, change_3p)] += 1
        detail.append({
            "miRNA_ID": row["miRNA_ID"],
            "Family": row["Family"],
            "miRNA_loci": row["miRNA_loci"],
            "Seq-ID": row["Seq-ID"],
            "Group_for_comparison": "isomiR with detected canonical" if canonical is not None else "isomiR without detected canonical",
            "Canonical_ID": canonical_id,
            "Canonical_source": canonical_source,
            "Fasta_ID_used": fasta_id,
            "Fasta_match_type": fasta_match_type,
            "Report_status": row["Report_status"],
            "Conservation": row["Conservation"],
            **comparison,
        })
    return counts, detail


def library_sets(seq_ids):
    sets = {seq_id: set() for seq_id in seq_ids}
    if not RAW_EXPRESSION_FILE.exists():
        return sets
    with RAW_EXPRESSION_FILE.open(encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        for row in reader:
            library_id, seq_id = row["sample"], row["Seq-ID"]
            if seq_id in sets:
                sets[seq_id].add(library_id)
    return sets


DATA_DIR.mkdir(parents=True, exist_ok=True)
FIGURE_DIR.mkdir(parents=True, exist_ok=True)
s2, expression, tissue_columns = read_workbook()
tissue_order_index = [tissue_columns.index(tissue) for tissue in PANEL_E_TISSUE_ORDER]

rich_family_rows = build_rich_family_summary(s2)
write_tsv(DATA_DIR / "isomiR_rich_family_summary.tsv", ["Family", "Reported", "Unreported", "isomiRs", "Total", "isomiR_percentage"], rich_family_rows)

pair_rows, no_canonical_rows = build_canonical_isomir_pair_tables(s2)
pair_fields = [
    "No.", "miRNA_ID", "Chr", "M_start", "M_end", "H_start", "H_end", "Strand",
    "Seq-ID", "Sequences_Mature", "Family", "miRNA_loci", "Hairpin_loci",
    "miRBase-mature", "miRBase-hairpin", "Report_status", "Conservation", "Group",
]
write_tsv(DATA_DIR / "canonical_with_isomiR_and_isomiR_mature_miRNAs.txt", pair_fields, pair_rows)
write_tsv(DATA_DIR / "isomiRs_without_detected_canonical_miRNA.txt", pair_fields + ["Expected_canonical_ID"], no_canonical_rows)

detection_pair_rows = build_detection_pair_rows(s2)
detection_pair_fields = [
    "Canonical_miRNA_ID",
    "Canonical_Seq_ID",
    "Canonical_library_detection_count",
    "IsomiR_miRNA_ID",
    "IsomiR_Seq_ID",
    "IsomiR_library_detection_count",
    "Canonical_detected",
    "Family",
    "miRNA_loci",
]
write_tsv(DATA_DIR / "canonical_isomiR_library_detection_pair_scatter.tsv", detection_pair_fields, detection_pair_rows)

det_rows = []
for row in pair_rows:
    det_rows.append({
        "miRNA_ID": row["miRNA_ID"],
        "Seq-ID": row["Seq-ID"],
        "Family": row["Family"],
        "miRNA_loci": row["miRNA_loci"],
        "Group": row["Group"],
        "Library_detection_count": library_count_from_seq_id(row["Seq-ID"]),
        "Detection_bin": detection_bin(library_count_from_seq_id(row["Seq-ID"])),
    })
write_tsv(DATA_DIR / "canonical_isomiR_library_detection_detail.tsv", list(det_rows[0]), det_rows)

groups = {
    "canonical": [r["Library_detection_count"] for r in det_rows if r["Group"] == "canonical"],
    "isomiR": [r["Library_detection_count"] for r in det_rows if r["Group"] == "isomiR"],
}
test = mannwhitneyu(groups["canonical"], groups["isomiR"], alternative="two-sided")
with (DATA_DIR / "canonical_isomiR_library_detection_summary.tsv").open("w", encoding="utf-8", newline="") as out:
    writer = csv.writer(out, delimiter="\t", lineterminator="\n")
    writer.writerow(["Metric", "Value"])
    writer.writerow(["input_rows", len(det_rows)])
    writer.writerow(["mann_whitney_u", f"{test.statistic:.6g}"])
    writer.writerow(["mann_whitney_p_value", f"{test.pvalue:.6g}"])
    writer.writerow([])
    writer.writerow(["Group", "n", "mean", "median", "q1", "q3", "min", "max"])
    for group in ["canonical", "isomiR"]:
        arr = np.asarray(groups[group], dtype=float)
        writer.writerow([group, len(arr), f"{arr.mean():.6f}", f"{np.median(arr):.6f}", f"{np.quantile(arr,0.25):.6f}", f"{np.quantile(arr,0.75):.6f}", int(arr.min()), int(arr.max())])
    writer.writerow([])
    writer.writerow(["Detection_bin", "canonical_count", "canonical_percentage", "isomiR_count", "isomiR_percentage"])
    for label in ["1", "2-5", "6-10", "11-50", "51-100", ">100"]:
        cc = sum(detection_bin(v) == label for v in groups["canonical"])
        ic = sum(detection_bin(v) == label for v in groups["isomiR"])
        writer.writerow([label, cc, f"{cc/len(groups['canonical'])*100:.2f}", ic, f"{ic/len(groups['isomiR'])*100:.2f}"])

terminal_change_counts, sequence_difference_rows = terminal_changes(pair_rows, no_canonical_rows, s2)
write_tsv(DATA_DIR / "isomiR_sequence_difference_detail.tsv", list(sequence_difference_rows[0]), sequence_difference_rows)
with (DATA_DIR / "isomiR_sequence_difference_summary.tsv").open("w", encoding="utf-8", newline="") as out:
    writer = csv.writer(out, delimiter="\t", lineterminator="\n")
    writer.writerow(["Section", "Metric", "Group", "Value"])
    writer.writerow(["overall", "isomiR_records_total", "all", len([r for r in s2 if r["Group"] == "isomiR"])])
    writer.writerow(["overall", "isomiR_records_compared", "all", len(sequence_difference_rows)])
    for (c5, c3), count in sorted(terminal_change_counts.items(), key=lambda kv: (-kv[1], kv[0])):
        writer.writerow(["terminal_change_combination", f"{c5},{c3}", "all", count])

plt.rcParams.update({
    "font.family": "Arial",
    "font.size": 11,
    "axes.linewidth": 1.1,
    "xtick.major.width": 1.1,
    "ytick.major.width": 1.1,
    "figure.facecolor": "white",
})

fig = plt.figure(figsize=(10.8, 13.2))
grid = fig.add_gridspec(3, 2, hspace=0.52, wspace=0.38)

# A. isomiR production by locus class and conservation (arm-unit level).
ax_a = fig.add_subplot(grid[0, 0])

# Merge known-family-new + novel-family-new → Non-reference
def merge_locus_class(lc: str) -> str:
    lc = str(lc).strip() if lc else ""
    if lc == "reference":
        return "Reference"
    return "Non-reference"

# Build arm-unit table: (miRNA_loci, arm_status) as unit
arm_data = defaultdict(lambda: {"locus_class": "", "conservation": set(), "isomiR": False})
for row in s2:
    locus = str(row.get("miRNA_loci") or "").strip()
    arm = str(row.get("arm_status") or "").strip()
    if not locus or not arm:
        continue
    key = (locus, arm)
    arm_data[key]["locus_class"] = merge_locus_class(row.get("locus_class", ""))
    arm_data[key]["conservation"].add(str(row.get("Conservation") or "").lower())
    if row["Group"] == "isomiR":
        arm_data[key]["isomiR"] = True

# Summarize by (merged_locus_class × Conservation)
order = [
    ("Reference",     "conserved"),
    ("Reference",     "specific"),
    ("Non-reference", "conserved"),
    ("Non-reference", "specific"),
]
a_summary = {k: {"total": 0, "isomiR": 0} for k in order}
for (_locus, _arm), rec in arm_data.items():
    cons = "conserved" if "conserved" in rec["conservation"] else "specific"
    lc = rec["locus_class"]
    key = (lc, cons)
    if key not in a_summary:
        continue
    a_summary[key]["total"] += 1
    if rec["isomiR"]:
        a_summary[key]["isomiR"] += 1

# Locus-level counts for bottom labels
locus_class_set = defaultdict(set)
for (_locus, _arm), rec in arm_data.items():
    locus_class_set[rec["locus_class"]].add(_locus)
ref_locus_n = len(locus_class_set.get("Reference", set()))
nonref_locus_n = len(locus_class_set.get("Non-reference", set()))

# Write TSV
write_tsv(
    DATA_DIR / "isomiR_production_by_locus_class_conservation.tsv",
    ["Locus_class", "Conservation", "total_arm_units", "isomiR_arm_units", "isomiR_percentage"],
    [
        {
            "Locus_class": lc,
            "Conservation": cons,
            "total_arm_units": a_summary[(lc, cons)]["total"],
            "isomiR_arm_units": a_summary[(lc, cons)]["isomiR"],
            "isomiR_percentage": f"{a_summary[(lc, cons)]['isomiR'] / a_summary[(lc, cons)]['total'] * 100:.1f}" if a_summary[(lc, cons)]["total"] else "0.0",
        }
        for lc, cons in order
    ],
)

bar_data = []
for lc, cons in order:
    t = a_summary[(lc, cons)]["total"]
    iso = a_summary[(lc, cons)]["isomiR"]
    pct = iso / t * 100 if t else 0
    bar_data.append({"pct": pct, "isomiR": iso, "total": t})

x = np.arange(len(bar_data))
pcts = [d["pct"] for d in bar_data]
totals_a = [d["total"] for d in bar_data]
iso_a = [d["isomiR"] for d in bar_data]
bar_colors = [BLUE if cons == "conserved" else ORANGE for _, cons in order]

ax_a.bar(x, pcts, color=bar_colors, edgecolor="white", linewidth=0.8, width=0.7)
for i, (p, iso, tot) in enumerate(zip(pcts, iso_a, totals_a)):
    ax_a.text(i, p + 1.2, f"{p:.1f}%\n({iso}/{tot})",
              ha="center", va="bottom", fontsize=10, fontweight="bold", color="#333333")

ax_a.set_xticks(x)
xt_labels_a = ax_a.set_xticklabels(["Conserved", "Specific"] * 2, fontsize=11)

# Place class labels first, then position horizontal lines midway
group_centers = [0.5, 2.5]
class_labels = ["Reference", "Non-reference"]
ref_texts_a = []
for center, label in zip(group_centers, class_labels):
    t = ax_a.text(center, -0.12, label, transform=ax_a.get_xaxis_transform(),
                  ha="center", va="top", fontsize=11, fontweight="bold")
    ref_texts_a.append(t)

fig.canvas.draw()
renderer = fig.canvas.get_renderer()

for i, center in enumerate(group_centers):
    left_bbox = xt_labels_a[2 * i].get_window_extent(renderer)
    right_bbox = xt_labels_a[2 * i + 1].get_window_extent(renderer)
    ref_bbox = ref_texts_a[i].get_window_extent(renderer)

    xtick_bottom_disp = min(left_bbox.y0, right_bbox.y0)
    xtick_bottom_axes = ax_a.transAxes.inverted().transform((0, xtick_bottom_disp))[1]
    ref_top_disp = ref_bbox.y1
    ref_top_axes = ax_a.transAxes.inverted().transform((0, ref_top_disp))[1]

    line_axes_y = (xtick_bottom_axes + ref_top_axes) / 2
    ref_texts_a[i].set_y(line_axes_y - 0.025)

    left_data = ax_a.transData.inverted().transform((left_bbox.x0, 0))[0]
    right_data = ax_a.transData.inverted().transform((right_bbox.x1, 0))[0]

    ax_a.plot([left_data, right_data], [line_axes_y, line_axes_y],
              transform=ax_a.get_xaxis_transform(), color="black", linewidth=0.8, clip_on=False)

ax_a.set_title("isomiR production by locus class", loc="left", fontsize=13, fontweight="bold", pad=9)
ax_a.set_ylabel("IsomiR-producing units (%)")
ax_a.set_ylim(0, max(pcts) * 1.18)
ax_a.yaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: f"{v:.0f}%"))
ax_a.spines[["top", "right"]].set_visible(False)
add_panel_label(fig, ax_a, "A")

# B. Top isomiR-rich families.
ax_b = fig.add_subplot(grid[0, 1])
plot_rows = list(reversed(rich_family_rows[:10]))
names = [r["Family"] for r in plot_rows]
reported = np.array([int(r["Reported"]) for r in plot_rows])
unreported = np.array([int(r["Unreported"]) for r in plot_rows])
isomirs_b = np.array([int(r["isomiRs"]) for r in plot_rows])
totals_b = reported + unreported + isomirs_b
ax_b.barh(names, reported, color=BLUE, label="reference")
ax_b.barh(names, unreported, left=reported, color=ORANGE, label="non-reference")
ax_b.barh(names, isomirs_b, left=reported + unreported, color=ISOMIR_BAR, label="isomiRs")
for i, row in enumerate(plot_rows):
    ax_b.text(reported[i] + unreported[i] + isomirs_b[i] / 2, i, str(row["isomiRs"]), va="center", ha="center", fontsize=10)
    ax_b.text(int(row["Total"]) + 0.45, i, str(row["Total"]), va="center", fontsize=10)
ax_b.set_title("Top 10 isomiR-rich families", loc="left", fontsize=13, fontweight="bold", pad=9)
ax_b.set_xlabel("Number of miRNA entries")
ax_b.set_ylabel("miRNA family")
ax_b.set_xlim(0, max(totals_b) * 1.16)
ax_b.legend(frameon=False, loc="lower right", fontsize=8.5)
style_axis(ax_b, grid_axis="x")
add_panel_label(fig, ax_b, "B")

# C. Terminal change combinations.
ax_c = fig.add_subplot(grid[1, 0])
change_bins = [-5] + list(range(-4, 5)) + [5]
change_labels = ["≤−5"] + [str(v).replace("-", "−") for v in range(-4, 5)] + ["≥5"]
heat = np.zeros((len(change_bins), len(change_bins)), dtype=int)

def change_bin_index(value):
    if value <= -5:
        return 0
    if value >= 5:
        return len(change_bins) - 1
    return change_bins.index(value)

for (change_5p, change_3p), count in terminal_change_counts.items():
    heat[change_bin_index(change_3p), change_bin_index(change_5p)] += count
image = ax_c.imshow(heat, cmap=C_HEATMAP_CMAP, aspect="auto")
ax_c.set_xticks(range(len(change_bins)), change_labels)
ax_c.set_yticks(range(len(change_bins)), change_labels)
ax_c.set_xlabel("5′ terminal change (nt)")
ax_c.set_ylabel("3′ terminal change (nt)")
ax_c.set_title("5′- and 3′-terminal shift combinations", loc="left", fontsize=13, fontweight="bold", pad=9)
for y in range(heat.shape[0]):
    for x0 in range(heat.shape[1]):
        if heat[y, x0] > 0:
            ax_c.text(x0, y, str(heat[y, x0]), ha="center", va="center", fontsize=9, color="black")
cbar = fig.colorbar(image, ax=ax_c, fraction=0.046, pad=0.03)
cbar.set_label("Number of isomiRs")
add_panel_label(fig, ax_c, "C")

# D. Canonical-isomiR library-detection pairs.
ax_d = fig.add_subplot(grid[1, 1])
d_x = np.asarray([row["Canonical_library_detection_count"] for row in detection_pair_rows], dtype=float)
d_y = np.asarray([row["IsomiR_library_detection_count"] for row in detection_pair_rows], dtype=float)
d_detected = np.asarray([row["Canonical_detected"] == "yes" for row in detection_pair_rows], dtype=bool)
d_detected_n = int(np.sum(d_detected))
d_undetected_n = int(np.sum(~d_detected))
ax_d.scatter(d_x[d_detected], d_y[d_detected], s=15, color=BLUE, alpha=0.7, edgecolors="none", label=f"canonical detected (n={d_detected_n})")
if np.any(~d_detected):
    ax_d.scatter(d_x[~d_detected], d_y[~d_detected], s=15, color=ORANGE, alpha=0.7, edgecolors="none", label=f"canonical not detected (n={d_undetected_n})")
d_upper = max(float(d_x.max()), float(d_y.max())) * 1.04
ax_d.plot([0, d_upper], [0, d_upper], color=GRAY, linewidth=0.9, linestyle="--", alpha=0.7)
ax_d.set_xlim(-15, d_upper)
ax_d.set_ylim(-15, d_upper)
ax_d.set_xlabel("Canonical library-detection count")
ax_d.set_ylabel("isomiR library-detection count")
ax_d.set_title("Canonical-isomiR library-detection pairs", loc="left", fontsize=13, fontweight="bold", pad=9)
ax_d.legend(frameon=False, fontsize=8.5, loc="upper right", handletextpad=0.25, labelspacing=0.35)
style_axis(ax_d)
add_panel_label(fig, ax_d, "D")

# E/F. Current gma-miR166g isomiR members.
mir166g_rows = sorted(
    [r for r in s2 if r["miRNA_loci"] == "gma-miR166g" and r["Group"] == "isomiR" and r["Seq-ID"] in expression],
    key=lambda r: r["miRNA_ID"],
)
mir166_ids = [r["Seq-ID"] for r in mir166g_rows]
mir166_labels = [r["miRNA_ID"].replace("gma-", "") for r in mir166g_rows]
mir166_library_counts = {
    "miR166g.v1": 433,
    "miR166g.v2": 214,
    "miR166g.v3": 138,
    "miR166g.v4": 123,
}
mir166_y_labels = [f"{label}\n({mir166_library_counts[label]})" for label in mir166_labels]
mir166_expression = np.vstack([expression[seq_id][tissue_order_index] for seq_id in mir166_ids])

ax_e = fig.add_subplot(grid[2, 0])
global_pearson, global_jaccard = read_global_pair_values()
label_x, label_y = read_miR166g_v3_v4_point()
ax_e.scatter(
    global_pearson,
    global_jaccard,
    s=15,
    color=BLUE,
    alpha=0.8,
    edgecolors="none",
    rasterized=True,
)
ax_e.scatter(
    [label_x],
    [label_y],
    s=130,
    facecolors="none",
    edgecolors="#E69F00",
    linewidths=1.4,
    zorder=5,
)
ax_e.annotate(
    "miR166g.v3 vs miR166g.v4",
    xy=(label_x, label_y),
    xytext=(-8, 8),
    textcoords="offset points",
    ha="right",
    va="bottom",
    fontsize=10,
)
ax_e.set_xlim(-1.05, 1.05)
ax_e.set_ylim(-0.03, 1.03)
ax_e.set_xlabel("Pearson correlation (r)")
ax_e.set_ylabel("Library-overlap Jaccard index")
ax_e.set_title("Expression correlation and library overlap", loc="left", fontsize=13, fontweight="bold", pad=9)
style_axis(ax_e)
add_panel_label(fig, ax_e, "E")

ax_f = fig.add_subplot(grid[2, 1])
n = len(mir166_ids)
mir166_correlation = np.eye(n)
mir166_p = np.zeros((n, n))
for i in range(n):
    for j in range(n):
        if i == j:
            continue
        result = pearsonr(mir166_expression[i], mir166_expression[j])
        mir166_correlation[i, j] = result.statistic
        mir166_p[i, j] = result.pvalue
libs = library_sets(mir166_ids)
mir166_jaccard = np.eye(n)
shared_n = np.zeros((n, n), dtype=int)
union_n = np.zeros((n, n), dtype=int)
for i, first in enumerate(mir166_ids):
    for j, second in enumerate(mir166_ids):
        shared = libs[first] & libs[second]
        union = libs[first] | libs[second]
        shared_n[i, j] = len(shared)
        union_n[i, j] = len(union)
        mir166_jaccard[i, j] = len(shared) / len(union) if union else np.nan
pearson_display = np.full_like(mir166_correlation, np.nan)
jaccard_display = np.full_like(mir166_jaccard, np.nan)
for i in range(n):
    for j in range(n):
        if i < j:
            pearson_display[i, j] = mir166_correlation[i, j]
        elif i > j:
            jaccard_display[i, j] = mir166_jaccard[i, j]
corr_image = ax_f.imshow(pearson_display, cmap=CORRELATION_CMAP, vmin=-1, vmax=1)
ax_f.imshow(jaccard_display, cmap=CORRELATION_CMAP, vmin=-1, vmax=1)
ax_f.set_xticks(range(n), mir166_labels, rotation=45, ha="right")
ax_f.set_yticks(range(n), mir166_y_labels)
for tick_label in ax_f.get_yticklabels():
    tick_label.set_multialignment("center")
    tick_label.set_linespacing(1.5)
ax_f.tick_params(length=0)
for i in range(n):
    for j in range(n):
        if i < j:
            value = mir166_correlation[i, j]
            ax_f.text(j, i, f"{value:.2f}{significance_label(mir166_p[i, j])}", ha="center", va="center", color="black", fontsize=9)
        elif i > j:
            value = mir166_jaccard[i, j]
            ax_f.text(j, i, f"{value:.3f}\n({shared_n[i,j]}/{union_n[i,j]})", ha="center", va="center", color="black", fontsize=9)
        else:
            ax_f.text(j, i, "1.00", ha="center", va="center", color="#444444", fontsize=9)
ax_f.set_xticks(np.arange(-0.5, n, 1), minor=True)
ax_f.set_yticks(np.arange(-0.5, n, 1), minor=True)
ax_f.grid(which="minor", color="white", linewidth=1.3)
ax_f.tick_params(which="minor", bottom=False, left=False)
pos_f = ax_f.get_position()
ax_f.set_position([pos_f.x0 - 0.045, pos_f.y0, pos_f.width, pos_f.height])
cbar_f = fig.colorbar(corr_image, ax=ax_f, fraction=0.046, pad=0.03)
cbar_f.set_ticks([-1, -0.5, 0, 0.5, 1])
cbar_f.set_ticklabels(["-1.0", "-0.5", "0.0", "0.5", "1.0"])
cbar_f.set_label("Correlation / similarity")
for axis_to_shift in (ax_f, cbar_f.ax):
    pos = axis_to_shift.get_position()
    axis_to_shift.set_position([pos.x0 - 0.031, pos.y0, pos.width, pos.height])
pos_d = ax_d.get_position()
pos_f = ax_f.get_position()
ax_f.set_title(
    "gma-miR166g isomiRs",
    x=(pos_d.x0 - pos_f.x0) / pos_f.width,
    ha="left",
    fontsize=13,
    fontweight="bold",
    pad=9,
)
add_panel_label(fig, ax_f, "F")

FIGURE_DIR.mkdir(parents=True, exist_ok=True)
FINAL_FIGURE_DIR.mkdir(parents=True, exist_ok=True)
fig.savefig(OUTPUT, dpi=600, bbox_inches="tight", facecolor="white")
fig.savefig(FINAL_OUTPUT, dpi=600, bbox_inches="tight", facecolor="white")
print(f"S2 rows={len(s2)}")
print(f"canonical={len(groups['canonical'])}")
print(f"isomiR={len(groups['isomiR'])}")
print(f"families={len(rich_family_rows)}")
print(f"miR166g_members={len(mir166_ids)}")
print(OUTPUT)
print(FINAL_OUTPUT)
