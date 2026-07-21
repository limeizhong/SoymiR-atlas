#!/usr/bin/env python3
from __future__ import annotations

import csv
from collections import Counter, defaultdict
from pathlib import Path

import pandas as pd


SCRIPT_DIR = Path(__file__).resolve().parent
MODULE_DIR = SCRIPT_DIR.parent
REPO_DIR = MODULE_DIR.parent
BASE = MODULE_DIR / "results"
FULL_TSV = BASE / "2814_precusor_miRNAs_annotated_precursor_overlap_unique_anchor_workflow.tsv"
SHORT_TSV = BASE / "2814_precusor_miRNAs_annotated_precursor_overlap_unique_anchor_workflow_short.tsv"
SIMPLIFIED_XLSX = BASE / "simplified_annotation.xlsx"
SIMPLIFIED_XLSX_ZH = BASE / "简化结果.xlsx"
SUPPLEMENTARY_XLSX = REPO_DIR / "05_Figures" / "input" / "source_tables" / "2026_merged_supplementary_tables.xlsx"

INTERMEDIATE_DIR = BASE / "intermediate"
COUNT_COMPARISON_TSV = INTERMEDIATE_DIR / "2814_annotation_category_rename_count_comparison.tsv"
QA_ISSUES_TSV = INTERMEDIATE_DIR / "2814_annotation_category_normalization_issues.tsv"
SUMMARY_TXT = BASE / "2814_precusor_miRNAs_annotated_precursor_overlap_unique_anchor_workflow_summary.txt"

CATEGORY_MAP = {
    "reported": "reference_matched",
    "known_member_variant": "reference_locus_variant",
    "unreported_mature_arm": "unannotated_opposite_arm_product",
    "unreported_mature_arm_variant": "unannotated_opposite_arm_variant",
    "known_family_new_member": "known_family_new_locus",
    "known_family_new_member_variant": "known_family_new_locus_variant",
    "new_family_member": "novel_family_new_locus",
    "new_family_member_variant": "novel_family_new_locus_variant",
}

CATEGORY_ORDER = [
    "reference_matched",
    "reference_locus_variant",
    "known_family_new_locus",
    "known_family_new_locus_variant",
    "novel_family_new_locus",
    "novel_family_new_locus_variant",
    "unannotated_opposite_arm_product",
    "unannotated_opposite_arm_variant",
]

NON_OPPOSITE_ATTRS = {
    "reference_matched": ("reference", "anchor", "anchor"),
    "reference_locus_variant": ("reference", "variant", "anchor"),
    "known_family_new_locus": ("known-family-new", "anchor", "anchor"),
    "known_family_new_locus_variant": ("known-family-new", "variant", "anchor"),
    "novel_family_new_locus": ("novel-family-new", "anchor", "anchor"),
    "novel_family_new_locus_variant": ("novel-family-new", "variant", "anchor"),
}

OPPOSITE_ATTRS = {
    "unannotated_opposite_arm_product": ("anchor", "opposite"),
    "unannotated_opposite_arm_variant": ("variant", "opposite"),
}

SHORT_COLUMNS = [
    "Seq-ID",
    "Sequences_Mature",
    "Chr",
    "M_start",
    "M_end",
    "H_start",
    "H_end",
    "Strand",
    "annotation_category_original",
    "annotation_category",
    "locus_class",
    "variant_status",
    "arm_status",
    "Annotation",
    "miRNA_Locus",
    "Reported_Status",
    "Conservation",
    "Conserved_Species_Count",
    "Source",
    "Family",
    "Position_Variant_Cluster",
    "CDHIT_Cluster_ID",
]

S2_COLUMNS = [
    "No.",
    "miRNA_ID",
    "Chr",
    "M_start",
    "M_end",
    "H_start",
    "H_end",
    "Strand",
    "Seq-ID",
    "annotation_category_original",
    "annotation_category",
    "locus_class",
    "variant_status",
    "arm_status",
    "Sequences_Mature",
    "Family",
    "miRNA_loci",
    "source",
    "Reported_Status",
    "Conservation",
    "Conserved_Species_Count",
    "Position_Variant_Cluster",
    "CDHIT_Cluster_ID",
]

S4_COLUMNS = [
    "Family",
    "Reference_matched_loci",
    "Nonreference_anchor_loci",
    "Variant_records",
    "Total",
    "Variant_percentage",
]

INT_COLUMNS = {"Chr", "M_start", "M_end", "H_start", "H_end", "Conserved_Species_Count"}


def normalize_int_like(value):
    if pd.isna(value):
        return ""
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return text


def read_table(path: Path) -> pd.DataFrame:
    df = pd.read_csv(path, sep="\t", dtype=str).fillna("")
    for col in INT_COLUMNS & set(df.columns):
        df[col] = df[col].map(normalize_int_like)
    return df


def old_category_column(df: pd.DataFrame) -> str:
    if "annotation_category_original" in df.columns:
        return "annotation_category_original"
    if "Status" in df.columns:
        return "Status"
    raise ValueError("Could not find Status or annotation_category_original column")


def precursor_key(row: pd.Series) -> tuple[str, str, str, str, str]:
    return (
        str(row.get("Chr", "")),
        str(row.get("H_start", "")),
        str(row.get("H_end", "")),
        str(row.get("Strand", "")),
        str(row.get("miRNA_Locus", row.get("miRNA_loci", ""))),
    )


def build_anchor_locus_classes(df: pd.DataFrame) -> tuple[dict[str, str], dict[tuple[str, str, str, str, str], str], list[dict[str, str]]]:
    by_locus: dict[str, set[str]] = defaultdict(set)
    by_precursor: dict[tuple[str, str, str, str, str], set[str]] = defaultdict(set)
    conflicts = []
    for _, row in df.iterrows():
        category = row["annotation_category"]
        if category not in NON_OPPOSITE_ATTRS:
            continue
        locus_class = NON_OPPOSITE_ATTRS[category][0]
        locus = str(row.get("miRNA_Locus", row.get("miRNA_loci", "")))
        if locus:
            by_locus[locus].add(locus_class)
        by_precursor[precursor_key(row)].add(locus_class)

    def collapse(mapping, key_name):
        out = {}
        for key, values in mapping.items():
            if len(values) == 1:
                out[key] = next(iter(values))
            else:
                conflicts.append({key_name: str(key), "issue_type": "anchor_locus_class_conflict", "values": ";".join(sorted(values))})
        return out

    return collapse(by_locus, "miRNA_Locus"), collapse(by_precursor, "precursor_key"), conflicts


def normalize_categories(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    df = df.copy()
    src_col = old_category_column(df)
    df["annotation_category_original"] = df[src_col]
    df["annotation_category"] = df["annotation_category_original"].map(CATEGORY_MAP).fillna("")

    issues: list[dict[str, str]] = []
    for idx, row in df[df["annotation_category"].eq("")].iterrows():
        issues.append(issue_row(row, "unrecognized_original_category", idx))

    locus_by_name, locus_by_precursor, conflicts = build_anchor_locus_classes(df)
    issues.extend(conflicts)

    locus_class = []
    variant_status = []
    arm_status = []

    for idx, row in df.iterrows():
        category = row["annotation_category"]
        if category in NON_OPPOSITE_ATTRS:
            lc, vs, arm = NON_OPPOSITE_ATTRS[category]
        elif category in OPPOSITE_ATTRS:
            vs, arm = OPPOSITE_ATTRS[category]
            locus = str(row.get("miRNA_Locus", row.get("miRNA_loci", "")))
            lc = locus_by_name.get(locus) or locus_by_precursor.get(precursor_key(row)) or "NA"
            if lc == "NA" and str(row.get("Source", "")).lower() in {"mirbase", "pmiren"}:
                lc = "reference"
            if lc == "NA":
                issues.append(issue_row(row, "opposite_arm_anchor_locus_class_unmatched", idx))
        else:
            lc, vs, arm = "NA", "NA", "NA"
        locus_class.append(lc)
        variant_status.append(vs)
        arm_status.append(arm)

    df["locus_class"] = locus_class
    df["variant_status"] = variant_status
    df["arm_status"] = arm_status

    issues.extend(validate_field_combinations(df))
    comparison = category_count_comparison(df)
    issue_df = pd.DataFrame(issues)
    if issue_df.empty:
        issue_df = pd.DataFrame(columns=["issue_type", "row_index", "Seq-ID", "Annotation", "miRNA_Locus", "annotation_category_original", "annotation_category"])
    return df, comparison, issue_df


def issue_row(row: pd.Series, issue_type: str, idx: int) -> dict[str, str]:
    return {
        "issue_type": issue_type,
        "row_index": str(idx + 1),
        "Seq-ID": str(row.get("Seq-ID", "")),
        "Annotation": str(row.get("Annotation", "")),
        "miRNA_Locus": str(row.get("miRNA_Locus", row.get("miRNA_loci", ""))),
        "annotation_category_original": str(row.get("annotation_category_original", row.get("Status", ""))),
        "annotation_category": str(row.get("annotation_category", "")),
    }


def validate_field_combinations(df: pd.DataFrame) -> list[dict[str, str]]:
    issues = []
    allowed_locus = {"reference", "known-family-new", "novel-family-new", "NA"}
    allowed_variant = {"anchor", "variant"}
    allowed_arm = {"anchor", "opposite"}
    reverse = {
        "reference_matched": ("reference", "anchor", "anchor"),
        "reference_locus_variant": ("reference", "variant", "anchor"),
        "known_family_new_locus": ("known-family-new", "anchor", "anchor"),
        "known_family_new_locus_variant": ("known-family-new", "variant", "anchor"),
        "novel_family_new_locus": ("novel-family-new", "anchor", "anchor"),
        "novel_family_new_locus_variant": ("novel-family-new", "variant", "anchor"),
    }
    for idx, row in df.iterrows():
        lc, vs, arm, cat = row["locus_class"], row["variant_status"], row["arm_status"], row["annotation_category"]
        if lc not in allowed_locus or vs not in allowed_variant or arm not in allowed_arm:
            issues.append(issue_row(row, "invalid_attribute_value", idx))
            continue
        if cat in reverse and (lc, vs, arm) != reverse[cat]:
            issues.append(issue_row(row, "category_attribute_conflict", idx))
        if cat in OPPOSITE_ATTRS and arm != "opposite":
            issues.append(issue_row(row, "opposite_arm_status_conflict", idx))
    return issues


def category_count_comparison(df: pd.DataFrame) -> pd.DataFrame:
    old_counts = Counter(df["annotation_category_original"])
    new_counts = Counter(df["annotation_category"])
    rows = []
    for old, new in CATEGORY_MAP.items():
        rows.append({
            "annotation_category_original": old,
            "original_count": old_counts.get(old, 0),
            "annotation_category": new,
            "normalized_count": new_counts.get(new, 0),
        })
    return pd.DataFrame(rows)


def reorder_columns(df: pd.DataFrame, short: bool) -> pd.DataFrame:
    drop_cols = ["Status"]
    df = df.drop(columns=[c for c in drop_cols if c in df.columns])
    preferred = SHORT_COLUMNS if short else [
        "File_Line",
        *SHORT_COLUMNS,
        "Evidence",
        "Matched_mature",
        "Matched_precursor",
        "Matched_chr",
        "Matched_start",
        "Matched_end",
        "Matched_strand",
        "Matched_precursor_overlap_bp",
    ]
    cols = [c for c in preferred if c in df.columns] + [c for c in df.columns if c not in preferred]
    return df[cols]


def write_tsv(df: pd.DataFrame, path: Path) -> None:
    df.to_csv(path, sep="\t", index=False, lineterminator="\n")


def write_xlsx(df: pd.DataFrame, path: Path) -> None:
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Sheet1")
        ws = writer.book["Sheet1"]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for col_cells in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max(max_len + 2, 10), 55)


def update_supplementary(short_df: pd.DataFrame) -> None:
    import openpyxl

    wb = openpyxl.load_workbook(SUPPLEMENTARY_XLSX)
    ws = wb["S2"]
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    ws.append(S2_COLUMNS)
    for idx, row in short_df.iterrows():
        ws.append([
            idx + 1,
            row["Annotation"],
            maybe_int(row["Chr"]),
            maybe_int(row["M_start"]),
            maybe_int(row["M_end"]),
            maybe_int(row["H_start"]),
            maybe_int(row["H_end"]),
            row["Strand"],
            row["Seq-ID"],
            row["annotation_category_original"],
            row["annotation_category"],
            row["locus_class"],
            row["variant_status"],
            row["arm_status"],
            row["Sequences_Mature"],
            row["Family"],
            row["miRNA_Locus"],
            row["Source"],
            row["Reported_Status"],
            row["Conservation"],
            maybe_int(row["Conserved_Species_Count"]),
            row["Position_Variant_Cluster"],
            row["CDHIT_Cluster_ID"],
        ])
    ws.freeze_panes = "A3"

    ws4 = wb["S4"]
    if ws4.max_row > 1:
        ws4.delete_rows(2, ws4.max_row - 1)
    ws4.append(S4_COLUMNS)
    for values in build_s4(short_df):
        ws4.append(values)
    ws4.freeze_panes = "A3"
    wb.save(SUPPLEMENTARY_XLSX)


def maybe_int(value):
    text = str(value)
    return int(text) if text.isdigit() else text


def build_s4(df: pd.DataFrame) -> list[list[object]]:
    rows = []
    for family, g in df.groupby("Family", sort=False):
        reference = int((g["annotation_category"] == "reference_matched").sum())
        variants = int((g["variant_status"] == "variant").sum())
        nonref_anchor = int(((g["variant_status"] == "anchor") & (g["annotation_category"] != "reference_matched")).sum())
        total = len(g)
        rows.append([family, reference, nonref_anchor, variants, total, round(variants / total * 100, 2) if total else 0])
    rows.sort(key=lambda x: (-x[4], str(x[0]).lower()))
    return rows


def write_summary(df: pd.DataFrame, issue_df: pd.DataFrame) -> None:
    lines = []
    lines.append("2814 precursor-location miRNA annotation summary\n")
    lines.append(f"Total records: {len(df)}\n")
    lines.append("\nannotation_category counts:\n")
    for cat in CATEGORY_ORDER:
        lines.append(f"{cat}: {int((df['annotation_category'] == cat).sum())}\n")
    lines.append("\nlocus_class counts:\n")
    for value, count in df["locus_class"].value_counts(dropna=False).items():
        lines.append(f"{value}: {count}\n")
    lines.append("\nvariant_status counts:\n")
    for value, count in df["variant_status"].value_counts(dropna=False).items():
        lines.append(f"{value}: {count}\n")
    lines.append("\narm_status counts:\n")
    for value, count in df["arm_status"].value_counts(dropna=False).items():
        lines.append(f"{value}: {count}\n")
    lines.append(f"\nUnassigned_or_conflict_records: {len(issue_df)}\n")
    SUMMARY_TXT.write_text("".join(lines), encoding="utf-8")


def main() -> None:
    INTERMEDIATE_DIR.mkdir(parents=True, exist_ok=True)
    full_raw = read_table(FULL_TSV)
    short_raw = read_table(SHORT_TSV)
    full_df, comparison, _ = normalize_categories(full_raw)
    short_df, _, issues = normalize_categories(short_raw)

    full_out = reorder_columns(full_df, short=False)
    short_out = reorder_columns(short_df, short=True)
    write_tsv(full_out, FULL_TSV)
    write_tsv(short_out, SHORT_TSV)
    write_xlsx(short_out, SIMPLIFIED_XLSX)
    write_xlsx(short_out, SIMPLIFIED_XLSX_ZH)
    update_supplementary(short_out)
    comparison.to_csv(COUNT_COMPARISON_TSV, sep="\t", index=False, lineterminator="\n")
    issues.to_csv(QA_ISSUES_TSV, sep="\t", index=False, lineterminator="\n")
    write_summary(short_out, issues)

    print(f"Total records: {len(short_out)}")
    print("annotation_category counts:")
    for cat in CATEGORY_ORDER:
        print(f"  {cat}: {int((short_out['annotation_category'] == cat).sum())}")
    print("locus_class counts:", dict(short_out["locus_class"].value_counts(dropna=False)))
    print("variant_status counts:", dict(short_out["variant_status"].value_counts(dropna=False)))
    print("arm_status counts:", dict(short_out["arm_status"].value_counts(dropna=False)))
    print(f"Unassigned_or_conflict_records: {len(issues)}")
    print(f"Wrote: {COUNT_COMPARISON_TSV}")
    print(f"Wrote: {QA_ISSUES_TSV}")


if __name__ == "__main__":
    main()
